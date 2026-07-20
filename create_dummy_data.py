"""
Creates a dummy Excel file with client certification data for testing.
"""
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

def create_dummy_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Certifications"

    # Header styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [
        "Client ID", "Full Name", "Company", "Email", "Phone (WhatsApp)",
        "Certification Name", "Certification ID", "Issue Date",
        "Expiry Date", "Renewal Link", "Status"
    ]

    ws.append(headers)
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    col_widths = [12, 20, 25, 32, 18, 30, 18, 14, 14, 40, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    today = datetime.today()

    # Dummy clients with varying expiry dates
    clients = [
        {
            "id": "CLT001", "name": "Rahul Sharma", "company": "TechCorp India Pvt Ltd",
            "email": "rahul.sharma@techcorp.in", "phone": "+919876543210",
            "cert": "ISO 9001:2015 Quality Management", "cert_id": "ISO-2021-4521",
            "issue": today - timedelta(days=365), "expiry": today + timedelta(days=7),
        },
        {
            "id": "CLT002", "name": "Priya Mehta", "company": "BuildRight Construction",
            "email": "priya.mehta@buildright.com", "phone": "+919812345678",
            "cert": "OSHA Safety Certification", "cert_id": "OSHA-2022-8834",
            "issue": today - timedelta(days=300), "expiry": today + timedelta(days=25),
        },
        {
            "id": "CLT003", "name": "Amit Verma", "company": "HealthFirst Pharma",
            "email": "amit.verma@healthfirst.co.in", "phone": "+919898765432",
            "cert": "GMP (Good Manufacturing Practice)", "cert_id": "GMP-2021-3367",
            "issue": today - timedelta(days=350), "expiry": today + timedelta(days=55),
        },
        {
            "id": "CLT004", "name": "Sneha Kapoor", "company": "EduTech Solutions",
            "email": "sneha.kapoor@edutech.in", "phone": "+919765432109",
            "cert": "ISO 27001 Information Security", "cert_id": "ISO27-2022-1192",
            "issue": today - timedelta(days=200), "expiry": today + timedelta(days=90),
        },
        {
            "id": "CLT005", "name": "Rajesh Nair", "company": "Logistics Plus",
            "email": "rajesh.nair@logisticsplus.com", "phone": "+919654321098",
            "cert": "HACCP Food Safety Certification", "cert_id": "HACCP-2020-5521",
            "issue": today - timedelta(days=400), "expiry": today - timedelta(days=5),  # Already expired
        },
        {
            "id": "CLT006", "name": "Anita Singh", "company": "GreenEnergy Corp",
            "email": "anita.singh@greenenergy.in", "phone": "+919543210987",
            "cert": "ISO 14001 Environmental Management", "cert_id": "ISO14-2023-7721",
            "issue": today - timedelta(days=100), "expiry": today + timedelta(days=180),
        },
        {
            "id": "CLT007", "name": "Vikram Joshi", "company": "AutoParts Manufacturing",
            "email": "vikram.joshi@autoparts.co.in", "phone": "+919432109876",
            "cert": "IATF 16949 Automotive Quality", "cert_id": "IATF-2022-9901",
            "issue": today - timedelta(days=330), "expiry": today + timedelta(days=30),
        },
        {
            "id": "CLT008", "name": "Deepika Rao", "company": "MedDevice India",
            "email": "deepika.rao@meddevice.in", "phone": "+919321098765",
            "cert": "ISO 13485 Medical Devices", "cert_id": "ISO13-2021-6634",
            "issue": today - timedelta(days=290), "expiry": today + timedelta(days=60),
        },
    ]

    row_fill_alt = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    renewal_base = "https://yourcertificationportal.com/renew?id="

    for i, c in enumerate(clients, 2):
        days_left = (c["expiry"] - today).days
        status = "EXPIRED" if days_left < 0 else ("CRITICAL" if days_left <= 7 else ("URGENT" if days_left <= 30 else ("DUE SOON" if days_left <= 60 else "ACTIVE")))

        row = [
            c["id"], c["name"], c["company"], c["email"], c["phone"],
            c["cert"], c["cert_id"],
            c["issue"].strftime("%d-%m-%Y"), c["expiry"].strftime("%d-%m-%Y"),
            renewal_base + c["cert_id"], status
        ]
        ws.append(row)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            if i % 2 == 0:
                cell.fill = row_fill_alt

        # Color-code status cell
        status_cell = ws.cell(row=i, column=11)
        status_colors = {
            "EXPIRED": "FF0000", "CRITICAL": "FF6600",
            "URGENT": "FFA500", "DUE SOON": "FFD700", "ACTIVE": "00B050"
        }
        status_cell.font = Font(bold=True, color=status_colors.get(status, "000000"))

    ws.row_dimensions[1].height = 25
    output_path = Path(__file__).parent / "clients_certifications.xlsx"
    wb.save(output_path)
    print(f"✅ Dummy Excel created: {output_path}")
    print(f"   {len(clients)} clients added with various expiry scenarios")

if __name__ == "__main__":
    create_dummy_excel()
