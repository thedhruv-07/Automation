"""One-off importer: converts a BIS ISI license register export into
clients_certifications.xlsx's single-sheet client roster schema.

Source sheets vary slightly in column set/order, so columns are looked up by
header name per sheet rather than by fixed position. Fields with no source
equivalent (Phone, Issue Date, Renewal Link) are left blank. Certification
Name is taken from each row's "Standard" column when the sheet has one (the
current single-sheet-with-all-standards-mixed template); if a sheet has no
"Standard" column at all -- the older one-sheet-per-standard layout -- or a
given row's Standard cell is blank, the sheet name is used instead (e.g.
"IS 269"), matching the original behavior. Certification ID and Client ID
both use the license number, since a license number is the natural unique
key in this dataset. Status is recomputed from Validity Date using this
project's own urgency thresholds -- the source "Status" column
(Operative/Cancelled/etc.) means something different and is not used.

Usage: python import_bis_isi_data.py "<path to source xlsx>"
"""
import sys
from datetime import datetime
from pathlib import Path

import openpyxl


class RowCollector:
    """Drop-in stand-in for an openpyxl worksheet's .append() — lets
    import_bis_isi_workbook() collect rows into a plain list instead of
    writing to a real worksheet, for callers that want the converted rows
    in memory (e.g. to merge with existing data) rather than a saved file."""

    def __init__(self):
        self.rows = []

    def append(self, row):
        self.rows.append(tuple(row))

OUTPUT_HEADERS = [
    "Client ID", "Full Name", "Company", "Email", "Phone (WhatsApp)",
    "Certification Name", "Certification ID", "Issue Date", "Expiry Date",
    "Renewal Link", "Status",
]

SOURCE_DATE_FORMATS = ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y")


def parse_validity_date(value):
    if isinstance(value, datetime):
        return value
    if value is None:
        return None
    for fmt in SOURCE_DATE_FORMATS:
        try:
            return datetime.strptime(str(value).strip(), fmt)
        except ValueError:
            continue
    return None


def compute_status(expiry_dt, today=None):
    today = today or datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
    days_left = (expiry_dt - today).days
    if days_left < 0:
        return "EXPIRED"
    if days_left <= 7:
        return "CRITICAL"
    if days_left <= 30:
        return "URGENT"
    if days_left <= 60:
        return "DUE SOON"
    return "ACTIVE"


def header_index_map(header_row):
    index_by_name = {}
    for i, name in enumerate(header_row):
        if name is None:
            continue
        index_by_name[str(name).strip().lower()] = i
    return index_by_name


def get(row, index_map, *names):
    for name in names:
        idx = index_map.get(name.lower())
        if idx is not None and idx < len(row):
            return row[idx]
    return None


def looks_like_bis_isi_workbook(wb) -> bool:
    """True if any sheet's header row has both a licence-no and firm-name column."""
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            header_row = next(ws.iter_rows(values_only=True))
        except StopIteration:
            continue
        index_map = header_index_map(header_row)
        if "licence no" in index_map and "firm name" in index_map:
            return True
    return False


def import_bis_isi_workbook(wb, out_ws, today=None):
    """Appends converted roster rows from an already-open BIS ISI workbook into out_ws."""
    today = today or datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)

    seen_client_ids = set()
    stats = {"sheets_used": 0, "sheets_empty": 0, "rows_written": 0, "rows_skipped_missing_key": 0}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            stats["sheets_empty"] += 1
            continue

        index_map = header_index_map(header_row)
        sheet_had_rows = False

        for row in rows_iter:
            if not row or row[0] is None:
                continue
            licence_no = get(row, index_map, "licence no")
            firm_name = get(row, index_map, "firm name")
            email = get(row, index_map, "email")
            validity_raw = get(row, index_map, "validity date", "validity")
            row_standard = get(row, index_map, "standard")
            cert_name = str(row_standard).strip() if row_standard else sheet_name.strip()

            if not licence_no or not validity_raw:
                stats["rows_skipped_missing_key"] += 1
                continue

            expiry_dt = parse_validity_date(validity_raw)
            if expiry_dt is None:
                stats["rows_skipped_missing_key"] += 1
                continue

            client_id = str(licence_no).strip()
            if client_id in seen_client_ids:
                client_id = f"{client_id}-{cert_name}"
            seen_client_ids.add(client_id)

            out_ws.append([
                client_id,
                firm_name,
                firm_name,
                email,
                None,
                cert_name,
                str(licence_no).strip(),
                None,
                expiry_dt.strftime("%d-%m-%Y"),
                None,
                compute_status(expiry_dt, today),
            ])
            stats["rows_written"] += 1
            sheet_had_rows = True

        if sheet_had_rows:
            stats["sheets_used"] += 1
        else:
            stats["sheets_empty"] += 1

    return stats


def import_bis_isi(source_path: str, output_path: str):
    wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.append(OUTPUT_HEADERS)

    stats = import_bis_isi_workbook(wb, out_ws)

    wb.close()
    out_wb.save(output_path)
    return stats


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python import_bis_isi_data.py \"<path to source xlsx>\"")
        sys.exit(1)

    source = sys.argv[1]
    output = str(Path(__file__).parent / "clients_certifications.xlsx")
    result = import_bis_isi(source, output)
    print("Sheets with data used:", result["sheets_used"])
    print("Sheets empty/skipped:", result["sheets_empty"])
    print("Rows written:", result["rows_written"])
    print("Rows skipped (missing licence no / validity date):", result["rows_skipped_missing_key"])
    print("Saved to:", output)
