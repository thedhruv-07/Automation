"""One-off/rerunnable correction: updates the cert_name (Certification Name)
of existing clients.db rows to match the "Standard" column in a corrected
BIS ISI master file, keyed by client_id (== Licence No).

Fixes rows that were imported with a wrong certification name by an earlier
BIS ISI import that fell back to a sheet name instead of reading a per-row
"Standard" column (see import_bis_isi_data.py's docstring for that history).

Only the cert_name field is touched, and only for rows whose client_id is
found in the source file AND whose current cert_name actually differs from
the source's Standard value. Every other field, and every client not present
in the source file (including every non-BIS-ISI client), is left untouched.

Usage: python fix_bis_isi_cert_names.py "<path to corrected master xlsx>" [--apply]
Without --apply, runs in dry-run mode: reports what WOULD change, writes nothing.
"""
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

from db import DEFAULT_DB_PATH, get_connection


def load_correct_standards(source_path) -> dict[str, str]:
    """Reads the first sheet with both a "Licence No" and "Standard" header
    into a {licence_no: standard} mapping."""
    wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                continue
            index = {str(h).strip().lower(): i for i, h in enumerate(header_row) if h is not None}
            if "licence no" not in index or "standard" not in index:
                continue
            lic_i, std_i = index["licence no"], index["standard"]
            mapping = {}
            for row in rows_iter:
                if not row or row[lic_i] is None or row[std_i] is None:
                    continue
                licence_no = str(row[lic_i]).strip()
                standard = str(row[std_i]).strip()
                if licence_no and standard:
                    mapping[licence_no] = standard
            if mapping:
                return mapping
        return {}
    finally:
        wb.close()


def find_cert_name_corrections(db_path, correct_standards: dict[str, str]) -> list[tuple[str, str, str]]:
    """Returns (client_id, old_cert_name, new_cert_name) for every roster row
    whose cert_name should change."""
    conn = get_connection(db_path)
    try:
        corrections = []
        for client_id, correct_standard in correct_standards.items():
            row = conn.execute(
                "SELECT cert_name FROM clients WHERE client_id = ?", (client_id,)
            ).fetchone()
            if row is None:
                continue
            if row["cert_name"] != correct_standard:
                corrections.append((client_id, row["cert_name"], correct_standard))
        return corrections
    finally:
        conn.close()


def apply_corrections(db_path, corrections: list[tuple[str, str, str]]) -> Path:
    """Backs up db_path, then writes the corrections. Returns the backup path."""
    backup_path = Path(db_path).parent / f"clients.backup-{datetime.now():%Y%m%d-%H%M%S}.db"
    shutil.copyfile(db_path, backup_path)

    conn = get_connection(db_path)
    try:
        conn.executemany(
            "UPDATE clients SET cert_name = ? WHERE client_id = ?",
            [(new, client_id) for client_id, _old, new in corrections],
        )
        conn.commit()
    finally:
        conn.close()
    return backup_path


def main():
    if len(sys.argv) < 2:
        print('Usage: python fix_bis_isi_cert_names.py "<path to corrected master xlsx>" [--apply]')
        sys.exit(1)

    source_path = sys.argv[1]
    apply = "--apply" in sys.argv[2:]

    correct_standards = load_correct_standards(source_path)
    print(f"Loaded {len(correct_standards)} licence -> standard mappings from {source_path}")

    corrections = find_cert_name_corrections(DEFAULT_DB_PATH, correct_standards)
    print(f"{len(corrections)} roster rows need a cert_name correction.")

    if not corrections:
        return

    print("Sample of corrections (first 10):")
    for client_id, old, new in corrections[:10]:
        print(f"  {client_id}: {old!r} -> {new!r}")

    if not apply:
        print("\nDry run only -- no changes written. Re-run with --apply to write these changes.")
        return

    backup_path = apply_corrections(DEFAULT_DB_PATH, corrections)
    print(f"Backed up previous database to {backup_path}")
    print(f"Applied {len(corrections)} corrections to {DEFAULT_DB_PATH}.")


if __name__ == "__main__":
    main()
