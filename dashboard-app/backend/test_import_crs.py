"""Tests for import_crs.py -- the BIS CRS registration export converter."""
import openpyxl

from import_crs import RowCollector, import_crs_workbook, looks_like_crs_workbook

PER_STANDARD_HEADERS = [
    "Application No.", "License No.", "Organization Name", "Country", "Indian Standard",
    "Product Name", "Product Category", "Grant Date", "Status", "E-mail", "Phone No.",
]


def _workbook(sheet_title, headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(headers)
    for row in rows:
        ws.append(row)
    return wb


def _multi_sheet_workbook(sheets):
    """sheets: list of (title, headers, rows) tuples -> one workbook with a
    sheet per tuple."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for title, headers, rows in sheets:
        ws = wb.create_sheet(title)
        ws.append(headers)
        for row in rows:
            ws.append(row)
    return wb


def test_looks_like_crs_workbook_detects_the_format():
    wb = _workbook("IS 13252", PER_STANDARD_HEADERS, [
        ["0846", "R-7019869", "Panache Digilife Limited", "India",
         "Is 13252(Part 1):2010/ Iec 60950-1: 2005", "Notebook", "Laptop/Notebook/Tablet",
         "2024-01-15", "Register", "jitendra.d@vardhamantechnology.com", "913322634755"],
    ])
    assert looks_like_crs_workbook(wb) is True


def test_looks_like_crs_workbook_does_not_match_bis_isi_headers():
    isi_headers = ["S. No.", "Licence No", "Firm Name", "Address", "District", "State",
                    "PIN Code", "Email", "Validity Date", "Status", "Variety", "Brand Names"]
    wb = _workbook("IS 999", isi_headers, [
        [1, "7777777", "Firm B", "Addr", "Dist", "State", "12345",
         "b@x.com", "2026-08-31", "Operative", "-", ""],
    ])
    assert looks_like_crs_workbook(wb) is False


def test_looks_like_crs_workbook_does_not_match_unrelated_workbook():
    wb = _workbook("Sheet1", ["Some", "Other", "Columns"], [["a", "b", "c"]])
    assert looks_like_crs_workbook(wb) is False


def test_import_crs_workbook_converts_rows_and_computes_expiry():
    wb = _workbook("IS 13252", PER_STANDARD_HEADERS, [
        ["0846", "R-7019869", "Panache Digilife Limited", "India",
         "Is 13252(Part 1):2010/ Iec 60950-1: 2005", "Notebook", "Laptop/Notebook/Tablet",
         "2024-01-15", "Register", "jitendra.d@vardhamantechnology.com", "913322634755"],
    ])
    collector = RowCollector()

    stats = import_crs_workbook(wb, collector)

    assert stats["rows_written"] == 1
    row = collector.rows[0]
    assert row[0] == "R-7019869"                                   # client_id
    assert row[1] == "Panache Digilife Limited"                    # name
    assert row[2] == "Panache Digilife Limited"                    # company
    assert row[3] == "jitendra.d@vardhamantechnology.com"          # email
    assert row[4] == "913322634755"                                # phone
    assert row[5] == "Is 13252(Part 1):2010/ Iec 60950-1: 2005"    # cert_name
    assert row[6] == "CRS"                                         # scheme
    assert row[7] == "R-7019869"                                   # cert_id
    assert row[8] == "15-01-2024"                                  # issue_date (Grant Date)
    assert row[9] == "15-01-2026"                                  # expiry_date (Grant Date + 2y)
    assert row[10] is None                                         # renewal_link
    assert row[11] in {"CRITICAL", "URGENT", "DUE SOON", "ACTIVE", "EXPIRED"}


def test_import_crs_workbook_reads_email_when_header_is_email_id():
    """The real BIS CRS export's header reads "E-mail Id", not the plain
    "E-mail" this file's own PER_STANDARD_HEADERS fixture uses -- that
    mismatch is exactly why production data came through with every
    email blank."""
    real_headers = [
        "Application No.", "License No.", "Organization Name", "Country", "Indian Standard",
        "Product Name", "Product Category", "Grant Date", "Status", "E-mail Id", "Phone No.",
    ]
    wb = _workbook("IS 13252", real_headers, [
        ["0846", "R-7019869", "Panache Digilife Limited", "India",
         "Is 13252(Part 1):2010/ Iec 60950-1: 2005", "Notebook", "Laptop/Notebook/Tablet",
         "2024-01-15", "Register", "jitendra.d@vardhamantechnology.com", "913322634755"],
    ])
    collector = RowCollector()

    import_crs_workbook(wb, collector)

    assert collector.rows[0][3] == "jitendra.d@vardhamantechnology.com"


def test_import_crs_workbook_skips_the_master_sheet():
    same_row = ["0846", "R-7019869", "Panache Digilife Limited", "India",
                "Is 13252(Part 1):2010/ Iec 60950-1: 2005", "Notebook", "Laptop/Notebook/Tablet",
                "2024-01-15", "Register", "jitendra.d@vardhamantechnology.com", "913322634755"]
    wb = _multi_sheet_workbook([
        ("BIS_CRS_Master", PER_STANDARD_HEADERS, [same_row]),
        ("IS 13252", PER_STANDARD_HEADERS, [same_row]),
    ])
    collector = RowCollector()

    stats = import_crs_workbook(wb, collector)

    assert stats["rows_written"] == 1
    assert len(collector.rows) == 1


def test_import_crs_workbook_master_sheet_name_match_is_case_insensitive():
    same_row = ["0846", "R-7019869", "Panache Digilife Limited", "India",
                "Is 13252(Part 1):2010/ Iec 60950-1: 2005", "Notebook", "Laptop/Notebook/Tablet",
                "2024-01-15", "Register", "jitendra.d@vardhamantechnology.com", "913322634755"]
    wb = _multi_sheet_workbook([
        ("bis_crs_master", PER_STANDARD_HEADERS, [same_row]),
        ("IS 13252", PER_STANDARD_HEADERS, [same_row]),
    ])
    collector = RowCollector()

    stats = import_crs_workbook(wb, collector)

    assert stats["rows_written"] == 1


def test_import_crs_workbook_imports_the_master_sheet_when_it_is_the_only_sheet():
    """Some real exports only ever contain the master rollup sheet, with no
    separate per-standard sheets alongside it -- in that case there's
    nothing to double-count, so the master sheet's rows must be imported
    rather than skipped."""
    same_row = ["0846", "R-7019869", "Panache Digilife Limited", "India",
                "Is 13252(Part 1):2010/ Iec 60950-1: 2005", "Notebook", "Laptop/Notebook/Tablet",
                "2024-01-15", "Register", "jitendra.d@vardhamantechnology.com", "913322634755"]
    wb = _workbook("BIS_CRS_Master", PER_STANDARD_HEADERS, [same_row])
    collector = RowCollector()

    stats = import_crs_workbook(wb, collector)

    assert stats["rows_written"] == 1
    assert len(collector.rows) == 1


def test_import_crs_workbook_skips_rows_missing_license_no_or_grant_date():
    wb = _workbook("IS 13252", PER_STANDARD_HEADERS, [
        [1, None, "No License Row", "India", "IS 13252", "Notebook", "Laptop", None,
         "Register", "a@x.com", "919000000000"],
        [2, "R-1111111", "No Grant Date Row", "India", "IS 13252", "Notebook", "Laptop", None,
         "Register", "b@x.com", "919000000001"],
    ])
    collector = RowCollector()

    stats = import_crs_workbook(wb, collector)

    assert stats["rows_written"] == 0
    assert stats["rows_skipped_missing_key"] == 2


def test_import_crs_workbook_dedups_same_license_no_across_different_standard_sheets():
    wb = _multi_sheet_workbook([
        ("IS 13252", PER_STANDARD_HEADERS, [
            ["1", "SAME123", "Firm A", "India", "IS 13252", "Notebook", "Laptop",
             "2024-01-15", "Register", "a@x.com", "919000000000"],
        ]),
        ("IS 616", PER_STANDARD_HEADERS, [
            ["2", "SAME123", "Firm A", "India", "IS 616", "Audio", "Speaker",
             "2024-01-15", "Register", "a@x.com", "919000000000"],
        ]),
    ])
    collector = RowCollector()

    stats = import_crs_workbook(wb, collector)

    assert stats["rows_written"] == 2
    client_ids = {row[0] for row in collector.rows}
    assert client_ids == {"SAME123", "SAME123-IS 616"}


def test_import_crs_workbook_collapses_same_license_and_standard_repeated_for_different_products():
    """Real CRS exports list one license/standard combo once per Product
    Category it covers -- e.g. the same license number + Indian Standard
    appearing three times, once per product category. These rows describe
    the same certification, not three different ones, so only the first
    should produce a roster row; the collision-suffix behavior (below) is
    only for a license number reused under a genuinely different standard."""
    wb = _workbook("IS 13252", PER_STANDARD_HEADERS, [
        ["0538", "R-4121916", "Corotronic Corporation", "Taiwan",
         "Is 13252(Part 1):2010/ Iec 60950-1: 2005", "X", "Automatic Data Processing Machine (Adpm)",
         "2024-01-15", "Register", "a@x.com", "919000000000"],
        ["0539", "R-4121916", "Corotronic Corporation", "Taiwan",
         "Is 13252(Part 1):2010/ Iec 60950-1: 2005", "Y", "Set Top Box",
         "2024-01-15", "Register", "a@x.com", "919000000000"],
        ["0540", "R-4121916", "Corotronic Corporation", "Taiwan",
         "Is 13252(Part 1):2010/ Iec 60950-1: 2005", "Z", "Tablet",
         "2024-01-15", "Register", "a@x.com", "919000000000"],
    ])
    collector = RowCollector()

    stats = import_crs_workbook(wb, collector)

    assert stats["rows_written"] == 1
    assert stats["rows_skipped_duplicate_product"] == 2
    client_ids = [row[0] for row in collector.rows]
    assert client_ids == ["R-4121916"]
    assert len(client_ids) == len(set(client_ids))


def test_import_crs_workbook_handles_feb_29_grant_date_in_non_leap_target_year():
    wb = _workbook("IS 13252", PER_STANDARD_HEADERS, [
        ["1", "R-2222222", "Firm C", "India", "IS 13252", "Notebook", "Laptop",
         "2024-02-29", "Register", "c@x.com", "919000000002"],
    ])
    collector = RowCollector()

    import_crs_workbook(wb, collector)

    assert collector.rows[0][9] == "28-02-2026"
