import io
import openpyxl
import main as main_module
from fastapi.testclient import TestClient
from main import app
from unittest.mock import patch
from db import upsert_clients, record_sent, load_sent_log, read_clients

client = TestClient(app)

HEADERS = [
    "Client ID", "Full Name", "Company", "Email", "Phone (WhatsApp)",
    "Certification Name", "Scheme", "Certification ID", "Issue Date", "Expiry Date",
    "Renewal Link", "Status",
]


def _write_xlsx(path, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)
    wb.save(path)


def _write_db(path, rows):
    upsert_clients(path, rows, mode="replace")


def test_health_endpoint():
    response = client.get("/api/health")
    assert response.status_code == 200
    assert response.json() == {"status": "ok"}


def test_resolve_allowed_origins_includes_localhost_by_default(monkeypatch):
    monkeypatch.delenv("DASHBOARD_ALLOWED_ORIGIN", raising=False)
    from main import _resolve_allowed_origins
    assert _resolve_allowed_origins() == ["http://localhost:5173", "http://127.0.0.1:5173"]


def test_resolve_allowed_origins_appends_configured_origin(monkeypatch):
    monkeypatch.setenv("DASHBOARD_ALLOWED_ORIGIN", "https://example.vercel.app")
    from main import _resolve_allowed_origins
    assert _resolve_allowed_origins() == [
        "http://localhost:5173", "http://127.0.0.1:5173", "https://example.vercel.app",
    ]


def test_protected_route_allows_request_when_auth_env_vars_unset(monkeypatch):
    monkeypatch.delenv("DASHBOARD_USERNAME", raising=False)
    monkeypatch.delenv("DASHBOARD_PASSWORD", raising=False)
    response = client.get("/api/stats")
    assert response.status_code == 200


def test_protected_route_rejects_missing_credentials_when_auth_configured(monkeypatch):
    monkeypatch.setenv("DASHBOARD_USERNAME", "admin")
    monkeypatch.setenv("DASHBOARD_PASSWORD", "s3cret")
    response = client.get("/api/stats")
    assert response.status_code == 401
    assert response.headers["www-authenticate"] == "Basic"


def test_protected_route_rejects_wrong_credentials_when_auth_configured(monkeypatch):
    monkeypatch.setenv("DASHBOARD_USERNAME", "admin")
    monkeypatch.setenv("DASHBOARD_PASSWORD", "s3cret")
    response = client.get("/api/stats", auth=("admin", "wrong"))
    assert response.status_code == 401


def test_protected_route_accepts_correct_credentials_when_auth_configured(monkeypatch):
    monkeypatch.setenv("DASHBOARD_USERNAME", "admin")
    monkeypatch.setenv("DASHBOARD_PASSWORD", "s3cret")
    response = client.get("/api/stats", auth=("admin", "s3cret"))
    assert response.status_code == 200


def test_health_endpoint_never_requires_auth(monkeypatch):
    monkeypatch.setenv("DASHBOARD_USERNAME", "admin")
    monkeypatch.setenv("DASHBOARD_PASSWORD", "s3cret")
    response = client.get("/api/health")
    assert response.status_code == 200


def test_get_clients_paginates_and_reports_total(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
        ["CLT004", "Sneha Kapoor", "EduTech", "s@x.com", "919765432109",
         "ISO 27001", "ISI", "ISO27-1", "01-01-2025", "15-10-2026", "https://x", "ACTIVE"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.setattr(main_module, "_today_str", lambda: "2026-07-18")

    response = client.get("/api/clients", params={"page": 1, "page_size": 1})
    assert response.status_code == 200
    data = response.json()
    assert data["total"] == 2
    assert data["page"] == 1
    assert len(data["rows"]) == 1


def test_get_clients_merges_alert_sent_today(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
        ["CLT004", "Sneha Kapoor", "EduTech", "s@x.com", "919765432109",
         "ISO 27001", "ISI", "ISO27-1", "01-01-2025", "15-10-2026", "https://x", "ACTIVE"],
    ])
    record_sent(db_path, "CLT001", "CRITICAL", "2026-07-18", "wamid.ABC", "919876543210", "2026-07-18T10:00:00")
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.setattr(main_module, "_today_str", lambda: "2026-07-18")

    response = client.get("/api/clients", params={"page_size": 50})
    data = response.json()["rows"]

    critical = next(r for r in data if r["client_id"] == "CLT001")
    assert critical["alert_sent_today"] is True
    active = next(r for r in data if r["client_id"] == "CLT004")
    assert active["alert_sent_today"] is None


def test_get_clients_filters_by_status_param(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT002", "Priya Mehta", "BuildRight", "p@x.com", "919812345678",
         "OSHA", "ISI", "OSHA-1", "01-01-2025", "11-08-2026", "https://x", "URGENT"],
        ["CLT004", "Sneha Kapoor", "EduTech", "s@x.com", "919765432109",
         "ISO 27001", "ISI", "ISO27-1", "01-01-2025", "15-10-2026", "https://x", "ACTIVE"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.setattr(main_module, "_today_str", lambda: "2026-07-18")

    response = client.get("/api/clients", params={"status": "URGENT", "page_size": 50})
    data = response.json()["rows"]
    assert len(data) == 1
    assert data[0]["client_id"] == "CLT002"


def test_get_stats_returns_counts_and_cert_types(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.setattr(main_module, "_today_str", lambda: "2026-07-18")

    response = client.get("/api/stats")
    assert response.status_code == 200
    data = response.json()
    assert data["status_counts"]["total"] == 1
    assert data["cert_types"] == ["ISO 9001"]


def test_eligible_count_returns_both_channels(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
        ["CLT002", "Priya Mehta", "BuildRight", "p@x.com", "919812345678",
         "OSHA", "ISI", "OSHA-1", "01-01-2025", "11-08-2026", "https://x", "URGENT"],
        ["CLT004", "Sneha Kapoor", "EduTech", "s@x.com", "919765432109",
         "ISO 27001", "ISI", "ISO27-1", "01-01-2025", "15-10-2026", "https://x", "ACTIVE"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.setattr(main_module, "_today_str", lambda: "2026-07-18")

    response = client.get("/api/eligible-count")
    assert response.status_code == 200
    assert response.json() == {"whatsapp": 2, "email": 2}


def test_eligible_count_filters_by_cert_type(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
        ["CLT002", "Priya Mehta", "BuildRight", "p@x.com", "919812345678",
         "OSHA", "ISI", "OSHA-1", "01-01-2025", "11-08-2026", "https://x", "URGENT"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.setattr(main_module, "_today_str", lambda: "2026-07-18")

    response = client.get("/api/eligible-count", params={"cert_type": "OSHA"})
    assert response.json() == {"whatsapp": 1, "email": 1}


def test_eligible_count_filters_by_search(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
        ["CLT002", "Priya Mehta", "BuildRight", "p@x.com", "919812345678",
         "OSHA", "ISI", "OSHA-1", "01-01-2025", "11-08-2026", "https://x", "URGENT"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.setattr(main_module, "_today_str", lambda: "2026-07-18")

    response = client.get("/api/eligible-count", params={"search": "BuildRight"})
    assert response.json() == {"whatsapp": 1, "email": 1}


def test_eligible_count_excludes_already_sent_today(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
    ])
    record_sent(db_path, "CLT001", "CRITICAL", "2026-07-18", "wamid.ABC", "919876543210", "2026-07-18T10:00:00")
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.setattr(main_module, "_today_str", lambda: "2026-07-18")

    response = client.get("/api/eligible-count")
    assert response.json() == {"whatsapp": 0, "email": 1}


def test_get_clients_filters_by_scheme_param(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
        ["CLT002", "Priya Mehta", "BuildRight", "p@x.com", "919812345678",
         "FMCS-Cert", "FMCS", "FMCS-1", "01-01-2025", "11-08-2026", "https://x", "URGENT"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.setattr(main_module, "_today_str", lambda: "2026-07-18")

    response = client.get("/api/clients", params={"scheme": "FMCS", "page_size": 50})
    data = response.json()["rows"]
    assert len(data) == 1
    assert data[0]["client_id"] == "CLT002"


def test_eligible_count_filters_by_scheme(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
        ["CLT002", "Priya Mehta", "BuildRight", "p@x.com", "919812345678",
         "FMCS-Cert", "FMCS", "FMCS-1", "01-01-2025", "11-08-2026", "https://x", "URGENT"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.setattr(main_module, "_today_str", lambda: "2026-07-18")

    response = client.get("/api/eligible-count", params={"scheme": "FMCS"})
    assert response.json() == {"whatsapp": 1, "email": 1}


def test_send_all_respects_scheme_filter(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
        ["CLT002", "Priya Mehta", "BuildRight", "p@x.com", "919812345678",
         "FMCS-Cert", "FMCS", "FMCS-1", "01-01-2025", "11-08-2026", "https://x", "URGENT"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.setenv("WHATSAPP_TOKEN", "tok")
    monkeypatch.setenv("PHONE_NUMBER_ID", "pid")
    monkeypatch.setenv("WHATSAPP_TEMPLATE_NAME_FMCS", "fmcs_renewal_alert")
    monkeypatch.setenv("WHATSAPP_TEMPLATE_LANG_FMCS", "en")

    mock_response = type("Resp", (), {
        "status_code": 200,
        "json": lambda self: {"messages": [{"id": "wamid.ABC"}]},
    })()
    with patch("whatsapp_renewal_alerts.requests.post", return_value=mock_response):
        response = client.post("/api/send-all", params={"scheme": "FMCS"})
        assert response.status_code == 200
        job_id = response.json()["job_id"]

        import time
        status_response = None
        for _ in range(50):
            status_response = client.get(f"/api/send-all/status/{job_id}")
            if status_response.json()["done"]:
                break
            time.sleep(0.05)

    final = status_response.json()
    assert final["done"] is True
    assert final["total"] == 1
    assert final["sent"] == 1


def test_export_clients_includes_scheme_column(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)

    response = client.get("/api/clients/export")
    assert response.status_code == 200
    lines = response.text.splitlines()
    assert lines[0].split(",")[6] == "Scheme"
    data_line = next(line for line in lines if "CLT001" in line)
    assert data_line.split(",")[6] == "ISI"


def test_send_all_respects_cert_type_filter(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
        ["CLT002", "Priya Mehta", "BuildRight", "p@x.com", "919812345678",
         "OSHA", "ISI", "OSHA-1", "01-01-2025", "11-08-2026", "https://x", "URGENT"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.setenv("WHATSAPP_TOKEN", "tok")
    monkeypatch.setenv("PHONE_NUMBER_ID", "pid")

    mock_response = type("Resp", (), {
        "status_code": 200,
        "json": lambda self: {"messages": [{"id": "wamid.ABC"}]},
    })()
    with patch("whatsapp_renewal_alerts.requests.post", return_value=mock_response):
        response = client.post("/api/send-all", params={"cert_type": "OSHA"})
        assert response.status_code == 200
        job_id = response.json()["job_id"]

        import time
        status_response = None
        for _ in range(50):
            status_response = client.get(f"/api/send-all/status/{job_id}")
            if status_response.json()["done"]:
                break
            time.sleep(0.05)

    final = status_response.json()
    assert final["done"] is True
    assert final["total"] == 1
    assert final["sent"] == 1


def test_send_all_respects_search_filter(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
        ["CLT002", "Priya Mehta", "BuildRight", "p@x.com", "919812345678",
         "OSHA", "ISI", "OSHA-1", "01-01-2025", "11-08-2026", "https://x", "URGENT"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.setenv("WHATSAPP_TOKEN", "tok")
    monkeypatch.setenv("PHONE_NUMBER_ID", "pid")

    mock_response = type("Resp", (), {
        "status_code": 200,
        "json": lambda self: {"messages": [{"id": "wamid.ABC"}]},
    })()
    with patch("whatsapp_renewal_alerts.requests.post", return_value=mock_response):
        response = client.post("/api/send-all", params={"search": "BuildRight"})
        assert response.status_code == 200
        job_id = response.json()["job_id"]

        import time
        status_response = None
        for _ in range(50):
            status_response = client.get(f"/api/send-all/status/{job_id}")
            if status_response.json()["done"]:
                break
            time.sleep(0.05)

    final = status_response.json()
    assert final["done"] is True
    assert final["total"] == 1
    assert final["sent"] == 1


def test_send_all_emails_respects_cert_type_filter(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
        ["CLT002", "Priya Mehta", "BuildRight", "p@x.com", "919812345678",
         "OSHA", "ISI", "OSHA-1", "01-01-2025", "11-08-2026", "https://x", "URGENT"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.setattr(main_module, "_today_str", lambda: "2026-07-18")
    monkeypatch.setenv("BREVO_API_KEY", "test-key")
    monkeypatch.setenv("EMAIL_SENDER", "sender@x.com")
    monkeypatch.delenv("DASHBOARD_TEST_EMAIL", raising=False)

    mock_response = type("Resp", (), {
        "status_code": 200,
        "json": lambda self: {"messageId": "brevo-1"},
    })()
    with patch("email_alerts.requests.post", return_value=mock_response):
        start_response = client.post("/api/send-all-emails", params={"cert_type": "OSHA"})
        assert start_response.status_code == 200
        job_id = start_response.json()["job_id"]

        import time
        status_response = None
        for _ in range(50):
            status_response = client.get(f"/api/send-all-emails/status/{job_id}")
            if status_response.json()["done"]:
                break
            time.sleep(0.05)

    final = status_response.json()
    assert final["done"] is True
    assert final["total"] == 1
    assert final["sent"] == 1


def test_send_all_emails_respects_search_filter(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
        ["CLT002", "Priya Mehta", "BuildRight", "p@x.com", "919812345678",
         "OSHA", "ISI", "OSHA-1", "01-01-2025", "11-08-2026", "https://x", "URGENT"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.setattr(main_module, "_today_str", lambda: "2026-07-18")
    monkeypatch.setenv("BREVO_API_KEY", "test-key")
    monkeypatch.setenv("EMAIL_SENDER", "sender@x.com")
    monkeypatch.delenv("DASHBOARD_TEST_EMAIL", raising=False)

    mock_response = type("Resp", (), {
        "status_code": 200,
        "json": lambda self: {"messageId": "brevo-1"},
    })()
    with patch("email_alerts.requests.post", return_value=mock_response):
        start_response = client.post("/api/send-all-emails", params={"search": "BuildRight"})
        assert start_response.status_code == 200
        job_id = start_response.json()["job_id"]

        import time
        status_response = None
        for _ in range(50):
            status_response = client.get(f"/api/send-all-emails/status/{job_id}")
            if status_response.json()["done"]:
                break
            time.sleep(0.05)

    final = status_response.json()
    assert final["done"] is True
    assert final["total"] == 1
    assert final["sent"] == 1


def test_export_clients_streams_csv_with_all_matching_rows(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
        ["CLT002", "Priya Mehta", "BuildRight", "p@x.com", "919812345678",
         "OSHA", "ISI", "OSHA-1", "01-01-2025", "11-08-2026", "https://x", "URGENT"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)

    response = client.get("/api/clients/export", params={"status": "CRITICAL"})
    assert response.status_code == 200
    assert "text/csv" in response.headers["content-type"]
    body = response.text
    assert "CLT001" in body
    assert "CLT002" not in body


def test_get_clients_rejects_out_of_range_pagination_params(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)

    assert client.get("/api/clients", params={"page": 0}).status_code == 422
    assert client.get("/api/clients", params={"page": -5}).status_code == 422
    assert client.get("/api/clients", params={"page_size": 501}).status_code == 422
    assert client.get("/api/clients", params={"page_size": 0}).status_code == 422


def test_get_clients_sort_dir_is_case_insensitive(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "B Name", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
        ["CLT002", "A Name", "BuildRight", "p@x.com", "919812345678",
         "OSHA", "ISI", "OSHA-1", "01-01-2025", "11-08-2026", "https://x", "URGENT"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)

    lower = client.get(
        "/api/clients", params={"sort_key": "name", "sort_dir": "desc", "page_size": 50},
    )
    upper = client.get(
        "/api/clients", params={"sort_key": "name", "sort_dir": "DESC", "page_size": 50},
    )
    mixed = client.get(
        "/api/clients", params={"sort_key": "name", "sort_dir": "Desc", "page_size": 50},
    )
    assert lower.status_code == upper.status_code == mixed.status_code == 200
    lower_ids = [r["client_id"] for r in lower.json()["rows"]]
    upper_ids = [r["client_id"] for r in upper.json()["rows"]]
    mixed_ids = [r["client_id"] for r in mixed.json()["rows"]]
    assert lower_ids == upper_ids == mixed_ids
    assert lower_ids == ["CLT001", "CLT002"]  # "B Name" sorts before "A Name" in descending order


def test_export_clients_escapes_leading_formula_characters(tmp_path, monkeypatch):
    """A company/name field starting with =, +, -, or @ must not be written
    verbatim to the CSV: Excel (and similar tools) would interpret it as a
    formula, which is a well-known CSV/formula-injection vector."""
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "=cmd|'/c calc'!A1", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)

    response = client.get("/api/clients/export")
    assert response.status_code == 200
    body = response.text
    data_line = next(line for line in body.splitlines() if "CLT001" in line)
    company_field = data_line.split(",")[2]
    assert company_field == "'=cmd|'/c calc'!A1"
    assert not company_field.startswith("=")


def test_email_preview_returns_subject_and_html(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)

    response = client.get("/api/email-preview/CLT001")
    assert response.status_code == 200
    data = response.json()
    assert data["subject"] == "Renew ISO 9001 — TechCorp"
    assert "Rahul Sharma" in data["html"]
    assert "Absolute Veritas" in data["html"]
    assert "24 July 2026" in data["html"]


def test_email_preview_uses_scheme_specific_content(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT004", "Deepa Rao", "FreshFoods", "d@x.com", "919000000001",
         "CRS-Cert", "CRS", "CRS-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.setenv("EMAIL_SUBJECT_TEMPLATE_CRS", "Registration renewal: {cert_name}")
    monkeypatch.setenv("EMAIL_INTRO_TEXT_CRS", "Your CRS registration for <strong>{company}</strong> needs renewal.")

    response = client.get("/api/email-preview/CLT004")

    assert response.status_code == 200
    data = response.json()
    assert data["subject"] == "Registration renewal: CRS-Cert"
    assert "Your CRS registration for <strong>FreshFoods</strong> needs renewal." in data["html"]


def test_email_preview_unknown_client_returns_404(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)

    response = client.get("/api/email-preview/NOPE")
    assert response.status_code == 404


def test_settings_info_reflects_env_and_never_exposes_token(monkeypatch):
    monkeypatch.setenv("WHATSAPP_TOKEN", "super-secret-token")
    monkeypatch.setenv("PHONE_NUMBER_ID", "pid123")
    monkeypatch.setenv("WHATSAPP_TEMPLATE_NAME", "cert_renewal_alert")
    monkeypatch.setenv("WHATSAPP_TEMPLATE_LANG", "en")

    response = client.get("/api/settings-info")
    assert response.status_code == 200
    data = response.json()
    assert data["template_name"] == "cert_renewal_alert"
    assert data["template_lang"] == "en"
    assert data["phone_number_id"] == "pid123"
    assert data["critical_days"] == 7
    assert data["urgent_days"] == 30
    assert "super-secret-token" not in response.text
    assert "token" not in data


def test_message_log_returns_entries_joined_with_client_data(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
    ])
    record_sent(db_path, "CLT001", "CRITICAL", "2026-07-18", "wamid.OLD", "919876543210", "2026-07-18T10:00:00")
    record_sent(db_path, "CLT001", "CRITICAL", "2026-07-19", "wamid.NEW", "919876543210", "2026-07-19T10:00:00")
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)

    response = client.get("/api/message-log")
    assert response.status_code == 200
    data = response.json()
    assert len(data) == 2
    # newest first
    assert data[0]["message_id"] == "wamid.NEW"
    assert data[0]["name"] == "Rahul Sharma"
    assert data[0]["company"] == "TechCorp"
    assert data[0]["status_tier"] == "CRITICAL"
    assert data[0]["phone"] == "919876543210"


def test_message_log_handles_client_no_longer_in_roster(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [])
    record_sent(db_path, "CLT999", "CRITICAL", "2026-07-18", "wamid.OLD", "919999999999", "2026-07-18T10:00:00")
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)

    response = client.get("/api/message-log")
    assert response.status_code == 200
    data = response.json()
    assert data[0]["name"] == "Unknown"


def _setup_one_client(tmp_path, monkeypatch, status="CRITICAL"):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", status],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.setattr(main_module, "_today_str", lambda: "2026-07-18")
    monkeypatch.setenv("WHATSAPP_TOKEN", "tok")
    monkeypatch.setenv("PHONE_NUMBER_ID", "pid123")
    monkeypatch.setenv("WHATSAPP_TEMPLATE_NAME", "cert_renewal_alert")
    monkeypatch.setenv("WHATSAPP_TEMPLATE_LANG", "en")
    monkeypatch.delenv("DASHBOARD_TEST_NUMBER", raising=False)
    return db_path


def test_send_alert_success(tmp_path, monkeypatch):
    db_path = _setup_one_client(tmp_path, monkeypatch)
    mock_response = type("Resp", (), {
        "status_code": 200,
        "json": lambda self: {"messages": [{"id": "wamid.ABC"}]},
    })()
    with patch("whatsapp_renewal_alerts.requests.post", return_value=mock_response):
        response = client.post("/api/send/CLT001")
    assert response.status_code == 200
    assert response.json() == {"status": "sent", "message_id": "wamid.ABC"}
    assert "CLT001|CRITICAL|2026-07-18" in load_sent_log(db_path)


def test_send_alert_unknown_client_returns_404(tmp_path, monkeypatch):
    _setup_one_client(tmp_path, monkeypatch)
    response = client.post("/api/send/NOPE")
    assert response.status_code == 404


def test_send_alert_ineligible_status_returns_400(tmp_path, monkeypatch):
    _setup_one_client(tmp_path, monkeypatch, status="ACTIVE")
    response = client.post("/api/send/CLT001")
    assert response.status_code == 400


def test_send_alert_duplicate_returns_409(tmp_path, monkeypatch):
    db_path = _setup_one_client(tmp_path, monkeypatch)
    record_sent(db_path, "CLT001", "CRITICAL", "2026-07-18", "wamid.OLD", None, "2026-07-18T09:00:00")
    response = client.post("/api/send/CLT001")
    assert response.status_code == 409


def test_send_alert_api_failure_returns_502(tmp_path, monkeypatch):
    _setup_one_client(tmp_path, monkeypatch)
    mock_response = type("Resp", (), {
        "status_code": 400,
        "json": lambda self: {"error": {"message": "Invalid parameter"}},
        "text": "Invalid parameter",
    })()
    with patch("whatsapp_renewal_alerts.requests.post", return_value=mock_response):
        response = client.post("/api/send/CLT001")
    assert response.status_code == 502
    assert "Invalid parameter" in response.json()["detail"]


def test_send_alert_uses_dashboard_test_number_override(tmp_path, monkeypatch):
    _setup_one_client(tmp_path, monkeypatch)
    monkeypatch.setenv("DASHBOARD_TEST_NUMBER", "919000000000")
    mock_response = type("Resp", (), {
        "status_code": 200,
        "json": lambda self: {"messages": [{"id": "wamid.TEST"}]},
    })()
    with patch("whatsapp_renewal_alerts.requests.post", return_value=mock_response) as mock_post:
        response = client.post("/api/send/CLT001")
    assert response.status_code == 200
    sent_payload = mock_post.call_args.kwargs["json"]
    assert sent_payload["to"] == "919000000000"


def test_send_alert_with_test_number_does_not_persist_dedup_log(tmp_path, monkeypatch):
    """Bug 1: a send redirected to DASHBOARD_TEST_NUMBER must NOT write the
    real client's dedup key to the sent_log table, or the real 9:30 AM CLI
    run (and future dashboard sends) would wrongly believe the client was
    already alerted today."""
    db_path = _setup_one_client(tmp_path, monkeypatch)
    original_log = load_sent_log(db_path)
    monkeypatch.setenv("DASHBOARD_TEST_NUMBER", "919000000000")
    mock_response = type("Resp", (), {
        "status_code": 200,
        "json": lambda self: {"messages": [{"id": "wamid.TEST"}]},
    })()
    with patch("whatsapp_renewal_alerts.requests.post", return_value=mock_response):
        response = client.post("/api/send/CLT001")
    assert response.status_code == 200
    assert response.json() == {"status": "sent", "message_id": "wamid.TEST"}

    # The sent_log table must be untouched: no dedup key was persisted.
    on_disk_log = load_sent_log(db_path)
    assert on_disk_log == original_log
    assert "CLT001|CRITICAL|2026-07-18" not in on_disk_log


def test_send_alert_concurrent_send_in_progress_returns_409(tmp_path, monkeypatch):
    """Bug 3: if a send for this client_id is already marked in-progress
    (e.g. an overlapping request got there first), a second request must be
    rejected with 409 rather than double-sending."""
    _setup_one_client(tmp_path, monkeypatch)
    main_module._pending_sends.add("CLT001")
    try:
        response = client.post("/api/send/CLT001")
    finally:
        main_module._pending_sends.discard("CLT001")
    assert response.status_code == 409
    assert "already in progress" in response.json()["detail"]


def test_send_alert_skipped_duplicate_from_send_one_alert_returns_409(tmp_path, monkeypatch):
    """Bug 2: if send_one_alert() itself reports skipped_duplicate (e.g. its
    internal dedup logic diverges from the endpoint's own pre-check in the
    future), the endpoint must still surface a 409, not a bogus 502."""
    _setup_one_client(tmp_path, monkeypatch)

    def fake_send_one_alert(*args, **kwargs):
        return {"action": "skipped_duplicate"}

    monkeypatch.setattr(main_module, "send_one_alert", fake_send_one_alert)
    response = client.post("/api/send/CLT001")
    assert response.status_code == 409
    assert "already sent today" in response.json()["detail"]


def test_send_alert_no_template_for_scheme_returns_400(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Priya Mehta", "BuildRight", "p@x.com", "919812345678",
         "CRS-Cert", "CRS", "CRS-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.setattr(main_module, "_today_str", lambda: "2026-07-18")
    monkeypatch.setenv("WHATSAPP_TOKEN", "tok")
    monkeypatch.setenv("PHONE_NUMBER_ID", "pid123")
    monkeypatch.delenv("WHATSAPP_TEMPLATE_NAME_CRS", raising=False)
    monkeypatch.delenv("WHATSAPP_TEMPLATE_LANG_CRS", raising=False)

    response = client.post("/api/send/CLT001")

    assert response.status_code == 400
    assert "CRS" in response.json()["detail"]


def test_send_all_starts_job_and_reports_progress(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.setenv("WHATSAPP_TOKEN", "tok")
    monkeypatch.setenv("PHONE_NUMBER_ID", "pid")

    # Note: whatsapp_renewal_alerts.run()/send_one_alert() default their
    # send_fn parameter to send_message at *def time*, so patching the
    # module-level `send_message` name (as one might expect) would not
    # actually intercept the call made from run()'s default argument.
    # Patching requests.post (which send_message calls by attribute lookup
    # at call time) is what actually takes effect, matching the pattern
    # already used by the /api/send/{id} tests above.
    mock_response = type("Resp", (), {
        "status_code": 200,
        "json": lambda self: {"messages": [{"id": "wamid.ABC"}]},
    })()
    with patch("whatsapp_renewal_alerts.requests.post", return_value=mock_response):
        response = client.post("/api/send-all")
        assert response.status_code == 200
        job_id = response.json()["job_id"]

        import time
        status_response = None
        for _ in range(50):
            status_response = client.get(f"/api/send-all/status/{job_id}")
            if status_response.json()["done"]:
                break
            time.sleep(0.05)

    final = status_response.json()
    assert final["done"] is True
    assert final["sent"] == 1
    assert final["total"] == 1


def test_send_all_reports_sent_for_all_alertable_statuses(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
        ["CLT002", "Priya Mehta", "BuildRight", "p@x.com", "919812345678",
         "OSHA", "ISI", "OSHA-1", "01-01-2025", "11-08-2026", "https://x", "URGENT"],
        ["CLT004", "Sneha Kapoor", "EduTech", "s@x.com", "919765432109",
         "ISO 27001", "ISI", "ISO27-1", "01-01-2025", "15-10-2026", "https://x", "ACTIVE"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.setenv("WHATSAPP_TOKEN", "tok")
    monkeypatch.setenv("PHONE_NUMBER_ID", "pid123")
    monkeypatch.delenv("DASHBOARD_TEST_NUMBER", raising=False)

    mock_response = type("Resp", (), {
        "status_code": 200,
        "json": lambda self: {"messages": [{"id": "wamid.ABC"}]},
    })()
    with patch("whatsapp_renewal_alerts.requests.post", return_value=mock_response):
        response = client.post("/api/send-all")
        job_id = response.json()["job_id"]

        import time
        status_response = None
        for _ in range(50):
            status_response = client.get(f"/api/send-all/status/{job_id}")
            if status_response.json()["done"]:
                break
            time.sleep(0.05)

    final = status_response.json()
    assert final["done"] is True
    assert final["total"] == 2  # only CRITICAL/URGENT are alertable; ACTIVE excluded
    assert final["sent"] == 2


def test_send_all_reports_skipped_no_template_for_unconfigured_scheme(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
        ["CLT002", "Priya Mehta", "BuildRight", "p@x.com", "919812345678",
         "CRS-Cert", "CRS", "CRS-1", "01-01-2025", "11-08-2026", "https://x", "URGENT"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.setenv("WHATSAPP_TOKEN", "tok")
    monkeypatch.setenv("PHONE_NUMBER_ID", "pid")
    monkeypatch.delenv("WHATSAPP_TEMPLATE_NAME_CRS", raising=False)
    monkeypatch.delenv("WHATSAPP_TEMPLATE_LANG_CRS", raising=False)

    mock_response = type("Resp", (), {
        "status_code": 200,
        "json": lambda self: {"messages": [{"id": "wamid.ABC"}]},
    })()
    with patch("whatsapp_renewal_alerts.requests.post", return_value=mock_response):
        response = client.post("/api/send-all")
        job_id = response.json()["job_id"]

        import time
        status_response = None
        for _ in range(50):
            status_response = client.get(f"/api/send-all/status/{job_id}")
            if status_response.json()["done"]:
                break
            time.sleep(0.05)

    final = status_response.json()
    assert final["done"] is True
    assert final["total"] == 2
    assert final["sent"] == 1
    assert final["skipped_no_template"] == 1


def test_send_all_uses_dashboard_test_number_override(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.setenv("WHATSAPP_TOKEN", "tok")
    monkeypatch.setenv("PHONE_NUMBER_ID", "pid123")
    monkeypatch.setenv("DASHBOARD_TEST_NUMBER", "919000000000")
    mock_response = type("Resp", (), {
        "status_code": 200,
        "json": lambda self: {"messages": [{"id": "wamid.TEST"}]},
    })()
    with patch("whatsapp_renewal_alerts.requests.post", return_value=mock_response) as mock_post:
        response = client.post("/api/send-all")
        job_id = response.json()["job_id"]

        import time
        status_response = None
        for _ in range(50):
            status_response = client.get(f"/api/send-all/status/{job_id}")
            if status_response.json()["done"]:
                break
            time.sleep(0.05)

    assert status_response.json()["done"] is True
    sent_payload = mock_post.call_args.kwargs["json"]
    assert sent_payload["to"] == "919000000000"


def test_send_all_job_reports_error_when_run_raises(tmp_path, monkeypatch):
    """If run() raises inside the background job (e.g. a locked DB or an
    unexpected error mid-send-loop), the job dict must record it so polling
    clients can distinguish a real crash from "0/0/0, nothing to do" -- not
    silently mark done=True with no trace of the failure."""
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.setenv("WHATSAPP_TOKEN", "tok")
    monkeypatch.setenv("PHONE_NUMBER_ID", "pid")

    def fake_run(*args, **kwargs):
        raise RuntimeError("database is locked")

    monkeypatch.setattr(main_module, "run", fake_run)

    response = client.post("/api/send-all")
    assert response.status_code == 200
    job_id = response.json()["job_id"]

    import time
    status_response = None
    for _ in range(50):
        status_response = client.get(f"/api/send-all/status/{job_id}")
        if status_response.json()["done"]:
            break
        time.sleep(0.05)

    final = status_response.json()
    assert final["done"] is True
    assert final["error"] is not None
    assert "database is locked" in final["error"]

    # The crash must not leave the bulk-in-progress flag stuck.
    assert main_module._bulk_in_progress is False


def test_send_all_job_error_is_none_on_success(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.setenv("WHATSAPP_TOKEN", "tok")
    monkeypatch.setenv("PHONE_NUMBER_ID", "pid")

    mock_response = type("Resp", (), {
        "status_code": 200,
        "json": lambda self: {"messages": [{"id": "wamid.ABC"}]},
    })()
    with patch("whatsapp_renewal_alerts.requests.post", return_value=mock_response):
        response = client.post("/api/send-all")
        job_id = response.json()["job_id"]

        import time
        status_response = None
        for _ in range(50):
            status_response = client.get(f"/api/send-all/status/{job_id}")
            if status_response.json()["done"]:
                break
            time.sleep(0.05)

    final = status_response.json()
    assert final["done"] is True
    assert final.get("error") is None


def test_send_all_status_returns_404_for_unknown_job(tmp_path, monkeypatch):
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", tmp_path / "clients.db")
    response = client.get("/api/send-all/status/does-not-exist")
    assert response.status_code == 404


def test_send_all_missing_env_var_resets_bulk_in_progress_flag(tmp_path, monkeypatch):
    """If required WHATSAPP_TOKEN/PHONE_NUMBER_ID env vars are missing, setup
    for the background job fails before the thread ever starts. _bulk_in_progress
    must still be reset in that case -- otherwise every future bulk AND
    per-client send would be permanently blocked with no way to recover short
    of restarting the server process (this was a regression introduced when
    send-all moved to a background job; the pre-migration synchronous
    implementation reset the flag in a finally around this exact lookup)."""
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.delenv("WHATSAPP_TOKEN", raising=False)
    monkeypatch.delenv("PHONE_NUMBER_ID", raising=False)

    response = client.post("/api/send-all")
    assert response.status_code == 500

    # The flag must have been reset: neither a subsequent bulk send nor a
    # subsequent per-client send should be blocked by a stuck flag left over
    # from the failed attempt.
    assert main_module._bulk_in_progress is False

    monkeypatch.setattr(main_module, "_today_str", lambda: "2026-07-18")
    monkeypatch.setenv("WHATSAPP_TOKEN", "tok")
    monkeypatch.setenv("PHONE_NUMBER_ID", "pid")
    mock_response = type("Resp", (), {
        "status_code": 200,
        "json": lambda self: {"messages": [{"id": "wamid.ABC"}]},
    })()
    with patch("whatsapp_renewal_alerts.requests.post", return_value=mock_response):
        second_response = client.post("/api/send/CLT001")
    assert second_response.status_code == 200


def test_send_all_alerts_blocks_concurrent_calls(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.setattr(main_module, "_bulk_in_progress", True)
    response = client.post("/api/send-all")
    assert response.status_code == 409


def test_send_alert_blocked_while_bulk_in_progress(tmp_path, monkeypatch):
    _setup_one_client(tmp_path, monkeypatch)
    monkeypatch.setattr(main_module, "_bulk_in_progress", True)
    response = client.post("/api/send/CLT001")
    assert response.status_code == 409


def test_send_all_alerts_blocked_while_per_client_send_in_progress(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    main_module._pending_sends.add("CLT999")
    try:
        response = client.post("/api/send-all")
        assert response.status_code == 409
    finally:
        main_module._pending_sends.discard("CLT999")


def test_client_template_returns_header_only_xlsx():
    response = client.get("/api/client-template")
    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in response.headers["content-disposition"]

    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows == [tuple(HEADERS)]


def test_upload_clients_success(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)

    upload_path = tmp_path / "upload.xlsx"
    _write_xlsx(upload_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
    ])

    with open(upload_path, "rb") as f:
        response = client.post(
            "/api/upload-clients",
            files={"file": ("clients.xlsx", f,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"import_format": "roster"},
        )
    assert response.status_code == 200
    assert response.json() == {
        "status": "ok", "row_count": 1, "format": "roster", "stats": {"rows_written": 1},
    }
    assert db_path.exists()
    assert read_clients(db_path)[0]["client_id"] == "CLT001"


def test_upload_clients_rejects_non_xlsx_extension(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    response = client.post(
        "/api/upload-clients",
        files={"file": ("clients.csv", b"not,a,real,xlsx", "text/csv")},
        data={"import_format": "roster"},
    )
    assert response.status_code == 400


def test_upload_clients_rejects_wrong_headers(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)

    upload_path = tmp_path / "bad.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Wrong", "Headers", "Here"])
    ws.append(["a", "b", "c"])
    wb.save(upload_path)

    with open(upload_path, "rb") as f:
        response = client.post(
            "/api/upload-clients",
            files={"file": ("bad.xlsx", f,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"import_format": "roster"},
        )
    assert response.status_code == 400
    assert not db_path.exists()


def test_upload_clients_rejects_empty_active_sheet_with_clear_message(tmp_path, monkeypatch):
    """A multi-sheet workbook where the active (last-selected) sheet has no
    rows must not crash with an unhandled StopIteration mislabeled as a
    generic invalid-file error."""
    db_path = tmp_path / "clients.db"
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)

    upload_path = tmp_path / "multi_sheet.xlsx"
    wb = openpyxl.Workbook()
    data_sheet = wb.active
    data_sheet.title = "Data"
    data_sheet.append(HEADERS)
    empty_sheet = wb.create_sheet("EmptyActive")
    wb.active = empty_sheet
    wb.save(upload_path)

    with open(upload_path, "rb") as f:
        response = client.post(
            "/api/upload-clients",
            files={"file": ("multi_sheet.xlsx", f,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"import_format": "roster"},
        )
    assert response.status_code == 400
    assert "EmptyActive" in response.json()["detail"]
    assert not db_path.exists()


def test_upload_clients_rejects_unknown_format(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)

    upload_path = tmp_path / "clients.xlsx"
    _write_xlsx(upload_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
    ])

    with open(upload_path, "rb") as f:
        response = client.post(
            "/api/upload-clients",
            files={"file": ("clients.xlsx", f,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"import_format": "fmcs"},
        )
    assert response.status_code == 400
    assert "fmcs" in response.json()["detail"].lower()


def test_upload_clients_missing_format_returns_422(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)

    upload_path = tmp_path / "clients.xlsx"
    _write_xlsx(upload_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
    ])

    with open(upload_path, "rb") as f:
        response = client.post(
            "/api/upload-clients",
            files={"file": ("clients.xlsx", f,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert response.status_code == 422


def test_upload_clients_selected_crs_but_file_is_roster_gives_targeted_error(tmp_path, monkeypatch):
    """Selecting the wrong format for a real file must name what the
    selected format actually needed, not the old generic
    "doesn't match any format" message -- proves the new per-format
    dispatch replaced the old cascade rather than just adding to it."""
    db_path = tmp_path / "clients.db"
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)

    upload_path = tmp_path / "clients.xlsx"
    _write_xlsx(upload_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
    ])

    with open(upload_path, "rb") as f:
        response = client.post(
            "/api/upload-clients",
            files={"file": ("clients.xlsx", f,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"import_format": "crs"},
        )
    assert response.status_code == 400
    detail = response.json()["detail"]
    assert "CRS" in detail
    assert "License No." in detail
    assert not db_path.exists()


def test_upload_clients_converts_raw_bis_isi_workbook(tmp_path, monkeypatch):
    """A raw BIS ISI licence export (govt column names, one sheet per IS
    standard, no Client ID/Phone/Company columns) should be auto-detected
    and converted into the roster schema, not rejected as a header mismatch."""
    db_path = tmp_path / "clients.db"
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)

    upload_path = tmp_path / "BIS ISI Data.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IS 302 (Part 2 Sec 30)"
    ws.append(["S. No.", "Licence No", "Firm Name", "Address", "District", "State",
               "PIN Code", "Email", "Validity Date", "Status", "Variety", "Brand Names"])
    ws.append([1, "9512485121", "Creative Hitech Private Limited", "Khasra No.-55, Rohad",
               "JHAJJAR", "HARYANA", "124501", "hod.quality@creativehitech.co.in",
               "24-07-2026", "Operative", "Fan", ""])
    wb.save(upload_path)

    with open(upload_path, "rb") as f:
        response = client.post(
            "/api/upload-clients",
            files={"file": ("BIS ISI Data.xlsx", f,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"import_format": "bis_isi"},
        )
    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "ok"
    assert body["format"] == "bis_isi"
    assert body["row_count"] == 1
    assert body["stats"]["rows_written"] == 1

    rows = read_clients(db_path)
    assert rows[0]["client_id"] == "9512485121"
    assert rows[0]["name"] == "Creative Hitech Private Limited"
    assert rows[0]["email"] == "hod.quality@creativehitech.co.in"
    assert rows[0]["cert_name"] == "IS 302 (Part 2 Sec 30)"


def test_upload_clients_converts_single_sheet_bis_isi_workbook_with_standard_column(tmp_path, monkeypatch):
    """The current BIS ISI export format: everything in one sheet, with a
    "Standard" column carrying what used to be the sheet name -- each row's
    certification should come from that column, not a shared sheet name."""
    db_path = tmp_path / "clients.db"
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)

    upload_path = tmp_path / "BIS_Final_Edited_Master_File.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["S. No.", "Standard", "Licence No", "Firm Name", "Address", "District", "State",
               "PIN Code", "Email", "Validity Date", "Status", "Variety", "Brand Names"])
    ws.append([1, "IS 1008", "4631257", "Prayagh Consumer Care Pvt. Unit II",
               "Survey No.527, 528 & 558 Chettanpally", "Mahbubnagar", "Telangana", "509316",
               "info@prayagh.com", "2026-08-31", "Operative", "-", ""])
    ws.append([2, "IS 10124 Part 2", "6298283", "Sudhakar PVC Products Private Limited",
               "Unit-2, Suryapet", "Suryapet", "Telangana", "508213",
               "sudhakarpipes@suryapet.com", "2026-12-31", "Operative", "Show Variety", ""])
    wb.save(upload_path)

    with open(upload_path, "rb") as f:
        response = client.post(
            "/api/upload-clients",
            files={"file": ("BIS_Final_Edited_Master_File.xlsx", f,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"import_format": "bis_isi"},
        )
    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "ok"
    assert body["format"] == "bis_isi"
    assert body["row_count"] == 2
    assert body["stats"]["rows_written"] == 2

    rows = {r["client_id"]: r for r in read_clients(db_path)}
    assert rows["4631257"]["cert_name"] == "IS 1008"
    assert rows["4631257"]["name"] == "Prayagh Consumer Care Pvt. Unit II"
    assert rows["6298283"]["cert_name"] == "IS 10124 Part 2"


def test_upload_clients_converts_raw_crs_workbook(tmp_path, monkeypatch):
    """A raw BIS CRS registration export (govt column names, one sheet per
    Indian Standard, no Client ID/Company columns) should be auto-detected
    and converted into the roster schema, not rejected as a header mismatch."""
    db_path = tmp_path / "clients.db"
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)

    upload_path = tmp_path / "CRS Data.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IS 13252"
    ws.append(["Application No.", "License No.", "Organization Name", "Country", "Indian Standard",
               "Product Name", "Product Category", "Grant Date", "Status", "E-mail", "Phone No."])
    ws.append(["0846", "R-7019869", "Panache Digilife Limited", "India",
               "Is 13252(Part 1):2010/ Iec 60950-1: 2005", "Notebook", "Laptop/Notebook/Tablet",
               "2024-01-15", "Register", "jitendra.d@vardhamantechnology.com", "913322634755"])
    wb.save(upload_path)

    with open(upload_path, "rb") as f:
        response = client.post(
            "/api/upload-clients",
            files={"file": ("CRS Data.xlsx", f,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"import_format": "crs"},
        )
    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "ok"
    assert body["format"] == "crs"
    assert body["row_count"] == 1
    assert body["stats"]["rows_written"] == 1

    rows = read_clients(db_path)
    assert rows[0]["client_id"] == "R-7019869"
    assert rows[0]["name"] == "Panache Digilife Limited"
    assert rows[0]["scheme"] == "CRS"
    assert rows[0]["cert_name"] == "Is 13252(Part 1):2010/ Iec 60950-1: 2005"


def test_upload_clients_backs_up_existing_file(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT999", "Old Client", "OldCo", "o@x.com", "919999999999",
         "Old Cert", "ISI", "OLD-1", "01-01-2025", "01-01-2026", "https://x", "ACTIVE"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)

    upload_path = tmp_path / "new.xlsx"
    _write_xlsx(upload_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
    ])

    with open(upload_path, "rb") as f:
        response = client.post(
            "/api/upload-clients",
            files={"file": ("new.xlsx", f,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"import_format": "roster"},
        )
    assert response.status_code == 200

    backup_path = db_path.parent / "clients.backup.db"
    assert backup_path.exists()


def test_upload_clients_rejects_blank_name_with_400_not_500(tmp_path, monkeypatch):
    """A blank name cell violates the `name TEXT NOT NULL` constraint in
    db.py, raising sqlite3.IntegrityError. The endpoint must surface this as
    a clear 400, not let it bubble up as an opaque 500, and must not leave
    the previously-existing roster wiped (mode="replace" deletes first)."""
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT999", "Old Client", "OldCo", "o@x.com", "919999999999",
         "Old Cert", "ISI", "OLD-1", "01-01-2025", "01-01-2026", "https://x", "ACTIVE"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)

    upload_path = tmp_path / "blank_name.xlsx"
    _write_xlsx(upload_path, [
        ["CLT001", None, "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
    ])

    with open(upload_path, "rb") as f:
        response = client.post(
            "/api/upload-clients",
            files={"file": ("blank_name.xlsx", f,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"import_format": "roster"},
        )
    assert response.status_code == 400
    assert "invalid" in response.json()["detail"].lower()

    # The pre-existing roster must survive untouched -- not wiped by the
    # DELETE that mode="replace" issues before the failed insert.
    rows = read_clients(db_path)
    assert {r["client_id"] for r in rows} == {"CLT999"}


def test_merge_clients_adds_new_and_keeps_existing(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT999", "Old Client", "OldCo", "o@x.com", "919999999999",
         "Old Cert", "ISI", "OLD-1", "01-01-2025", "01-01-2026", "https://x", "ACTIVE"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)

    upload_path = tmp_path / "new.xlsx"
    _write_xlsx(upload_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
    ])

    with open(upload_path, "rb") as f:
        response = client.post(
            "/api/merge-clients",
            files={"file": ("new.xlsx", f,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"import_format": "roster"},
        )
    assert response.status_code == 200
    body = response.json()
    assert body == {
        "status": "ok", "row_count": 2, "added": 1, "skipped_duplicates": 0,
        "format": "roster", "stats": {"rows_written": 1},
    }

    rows = read_clients(db_path)
    client_ids = {row["client_id"] for row in rows}
    assert client_ids == {"CLT999", "CLT001"}


def test_merge_clients_skips_duplicate_client_ids(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma (old data)", "TechCorp", "old@x.com", "919999999999",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "01-01-2026", "https://x", "ACTIVE"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)

    upload_path = tmp_path / "new.xlsx"
    _write_xlsx(upload_path, [
        ["CLT001", "Rahul Sharma (new data)", "TechCorp", "new@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
        ["CLT002", "Priya Mehta", "BuildRight", "p@x.com", "919812345678",
         "OSHA", "ISI", "OSHA-1", "01-01-2025", "11-08-2026", "https://x", "URGENT"],
    ])

    with open(upload_path, "rb") as f:
        response = client.post(
            "/api/merge-clients",
            files={"file": ("new.xlsx", f,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"import_format": "roster"},
        )
    assert response.status_code == 200
    body = response.json()
    assert body["row_count"] == 2
    assert body["added"] == 1
    assert body["skipped_duplicates"] == 1

    rows = read_clients(db_path)
    by_id = {row["client_id"]: row for row in rows}
    assert by_id["CLT001"]["name"] == "Rahul Sharma (old data)"  # old data kept, not overwritten
    assert "CLT002" in by_id


def test_merge_clients_converts_and_merges_raw_bis_isi_workbook(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["9512485121", "Existing Firm Name", "Existing Firm Name", "existing@x.com", None,
         "IS 302 (Part 2 Sec 30)", "ISI", "9512485121", None, "01-01-2026", None, "ACTIVE"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)

    upload_path = tmp_path / "BIS ISI Data.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IS 302 (Part 2 Sec 30)"
    ws.append(["S. No.", "Licence No", "Firm Name", "Email", "Validity Date"])
    ws.append([1, "9512485121", "Creative Hitech Private Limited", "hod@creativehitech.co.in", "24-07-2026"])
    ws.append([2, "8700138914", "Power Fan Industry", "bharatbijleeudhyog@gmail.com", "26-07-2026"])
    wb.save(upload_path)

    with open(upload_path, "rb") as f:
        response = client.post(
            "/api/merge-clients",
            files={"file": ("BIS ISI Data.xlsx", f,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"import_format": "bis_isi"},
        )
    assert response.status_code == 200
    body = response.json()
    assert body["format"] == "bis_isi"
    assert body["added"] == 1
    assert body["skipped_duplicates"] == 1
    assert body["row_count"] == 2

    rows = read_clients(db_path)
    by_id = {row["client_id"]: row for row in rows}
    assert by_id["9512485121"]["name"] == "Existing Firm Name"  # kept, not overwritten by upload
    assert by_id["8700138914"]["name"] == "Power Fan Industry"  # new client added


def test_merge_clients_into_empty_roster(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)

    upload_path = tmp_path / "new.xlsx"
    _write_xlsx(upload_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
    ])

    with open(upload_path, "rb") as f:
        response = client.post(
            "/api/merge-clients",
            files={"file": ("new.xlsx", f,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"import_format": "roster"},
        )
    assert response.status_code == 200
    body = response.json()
    assert body["row_count"] == 1
    assert body["added"] == 1
    assert body["skipped_duplicates"] == 0


def test_merge_clients_rejects_blank_name_with_400_and_rolls_back_batch(tmp_path, monkeypatch):
    """A blank name cell anywhere in the uploaded batch must surface as a
    400, and none of the batch's rows -- including a valid row that sorted
    before the bad one -- may be partially merged in (all-or-nothing)."""
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT999", "Old Client", "OldCo", "o@x.com", "919999999999",
         "Old Cert", "ISI", "OLD-1", "01-01-2025", "01-01-2026", "https://x", "ACTIVE"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)

    upload_path = tmp_path / "new.xlsx"
    _write_xlsx(upload_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
        ["CLT002", None, "BuildRight", "p@x.com", "919812345678",
         "OSHA", "ISI", "OSHA-1", "01-01-2025", "11-08-2026", "https://x", "URGENT"],
    ])

    with open(upload_path, "rb") as f:
        response = client.post(
            "/api/merge-clients",
            files={"file": ("new.xlsx", f,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"import_format": "roster"},
        )
    assert response.status_code == 400
    assert "invalid" in response.json()["detail"].lower()

    # CLT001 (valid, ordered before the bad row) must NOT have been
    # partially merged in -- only the pre-existing CLT999 remains.
    rows = read_clients(db_path)
    assert {r["client_id"] for r in rows} == {"CLT999"}


def test_merge_clients_rejects_non_xlsx_extension(tmp_path, monkeypatch):
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", tmp_path / "clients.db")
    response = client.post(
        "/api/merge-clients",
        files={"file": ("data.csv", b"not,a,spreadsheet", "text/csv")},
        data={"import_format": "roster"},
    )
    assert response.status_code == 400


def test_merge_clients_selected_bis_isi_but_file_is_roster_gives_targeted_error(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)

    upload_path = tmp_path / "clients.xlsx"
    _write_xlsx(upload_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
    ])

    with open(upload_path, "rb") as f:
        response = client.post(
            "/api/merge-clients",
            files={"file": ("clients.xlsx", f,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"import_format": "bis_isi"},
        )
    assert response.status_code == 400
    detail = response.json()["detail"]
    assert "BIS ISI" in detail
    assert "Firm Name" in detail


def test_merge_clients_backs_up_existing_file(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT999", "Old Client", "OldCo", "o@x.com", "919999999999",
         "Old Cert", "ISI", "OLD-1", "01-01-2025", "01-01-2026", "https://x", "ACTIVE"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)

    upload_path = tmp_path / "new.xlsx"
    _write_xlsx(upload_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
    ])

    with open(upload_path, "rb") as f:
        response = client.post(
            "/api/merge-clients",
            files={"file": ("new.xlsx", f,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"import_format": "roster"},
        )
    assert response.status_code == 200

    backup_path = db_path.parent / "clients.backup.db"
    assert backup_path.exists()


from db import load_email_sent_log, save_email_sent_log


def _setup_one_email_client(tmp_path, monkeypatch, status="CRITICAL"):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", status],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.setattr(main_module, "_today_str", lambda: "2026-07-18")
    monkeypatch.setenv("BREVO_API_KEY", "test-key")
    monkeypatch.setenv("EMAIL_SENDER", "sender@x.com")
    monkeypatch.delenv("DASHBOARD_TEST_EMAIL", raising=False)
    return db_path


def test_send_email_success(tmp_path, monkeypatch):
    db_path = _setup_one_email_client(tmp_path, monkeypatch)
    mock_response = type("Resp", (), {
        "status_code": 200,
        "json": lambda self: {"messageId": "brevo-1"},
    })()
    with patch("email_alerts.requests.post", return_value=mock_response):
        response = client.post("/api/send-email/CLT001")
    assert response.status_code == 200
    assert response.json() == {"status": "sent", "message_id": "brevo-1"}
    log = load_email_sent_log(db_path)
    assert "CLT001|CRITICAL|2026-07-18" in log


def test_send_email_unknown_client_returns_404(tmp_path, monkeypatch):
    _setup_one_email_client(tmp_path, monkeypatch)
    response = client.post("/api/send-email/NOPE")
    assert response.status_code == 404


def test_send_email_ineligible_status_returns_400(tmp_path, monkeypatch):
    _setup_one_email_client(tmp_path, monkeypatch, status="ACTIVE")
    response = client.post("/api/send-email/CLT001")
    assert response.status_code == 400


def test_send_email_no_email_on_file_returns_400(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT005", "No Email Co", "No Email Co", None, "919000000000",
         "ISO 9001", "ISI", "ISO-5", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.setenv("BREVO_API_KEY", "test-key")
    monkeypatch.setenv("EMAIL_SENDER", "sender@x.com")
    response = client.post("/api/send-email/CLT005")
    assert response.status_code == 400
    assert "email" in response.json()["detail"].lower()


def test_send_email_duplicate_returns_409(tmp_path, monkeypatch):
    db_path = _setup_one_email_client(tmp_path, monkeypatch)
    save_email_sent_log(db_path, {
        "CLT001|CRITICAL|2026-07-18": {"sent_at": "x", "message_id": "y", "email": "r@x.com"},
    })
    response = client.post("/api/send-email/CLT001")
    assert response.status_code == 409


def test_send_all_emails_starts_job_and_reports_progress(tmp_path, monkeypatch):
    _setup_one_email_client(tmp_path, monkeypatch)
    mock_response = type("Resp", (), {
        "status_code": 200,
        "json": lambda self: {"messageId": "brevo-1"},
    })()
    with patch("email_alerts.requests.post", return_value=mock_response):
        start_response = client.post("/api/send-all-emails")
        assert start_response.status_code == 200
        job_id = start_response.json()["job_id"]

        import time
        status_response = None
        for _ in range(50):
            status_response = client.get(f"/api/send-all-emails/status/{job_id}")
            if status_response.json()["done"]:
                break
            time.sleep(0.05)

    final = status_response.json()
    assert final["done"] is True
    assert final["sent"] == 1
    assert final["total"] == 1


def test_send_all_emails_status_returns_404_for_unknown_job():
    response = client.get("/api/send-all-emails/status/does-not-exist")
    assert response.status_code == 404


def test_send_all_emails_blocks_concurrent_calls(tmp_path, monkeypatch):
    _setup_one_email_client(tmp_path, monkeypatch)
    monkeypatch.setattr(main_module, "_email_bulk_in_progress", True)
    response = client.post("/api/send-all-emails")
    assert response.status_code == 409


def test_send_all_emails_does_not_block_on_whatsapp_bulk_in_progress(tmp_path, monkeypatch):
    """The two channels' bulk-send locks must be independent."""
    _setup_one_email_client(tmp_path, monkeypatch)
    monkeypatch.setattr(main_module, "_bulk_in_progress", True)  # WhatsApp's flag, not email's
    mock_response = type("Resp", (), {
        "status_code": 200,
        "json": lambda self: {"messageId": "brevo-1"},
    })()
    with patch("email_alerts.requests.post", return_value=mock_response):
        response = client.post("/api/send-all-emails")
    assert response.status_code == 200


def test_notices_list_includes_transition_facilitation_2026():
    response = client.get("/api/notices")
    assert response.status_code == 200
    ids = {n["id"] for n in response.json()}
    assert "transition_facilitation_2026" in ids


def test_notice_eligible_count_unknown_notice_returns_404(tmp_path, monkeypatch):
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", tmp_path / "clients.db")
    response = client.get("/api/notices/does_not_exist/eligible-count")
    assert response.status_code == 404


def test_notice_eligible_count_reflects_scheme_filter(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "OSHA", "CRS", "OSHA-1", "01-01-2025", "24-07-2026", "https://x", "CRITICAL"],
        ["CLT002", "Priya Mehta", "BuildRight", "p@x.com", "919812345678",
         "ISO 9001", "ISI", "ISO-1", "01-01-2025", "24-07-2026", "https://x", "ACTIVE"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)

    response = client.get("/api/notices/transition_facilitation_2026/eligible-count", params={"scheme": "CRS"})
    assert response.status_code == 200
    assert response.json() == {"whatsapp": 1, "email": 1}


def test_send_notice_whatsapp_unknown_notice_returns_404(tmp_path, monkeypatch):
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", tmp_path / "clients.db")
    response = client.post("/api/notices/does_not_exist/send-whatsapp")
    assert response.status_code == 404


def test_send_notice_whatsapp_starts_job_and_reports_progress(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "OSHA", "CRS", "OSHA-1", "01-01-2025", "24-07-2026", "https://x", "ACTIVE"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.setenv("WHATSAPP_TOKEN", "tok")
    monkeypatch.setenv("PHONE_NUMBER_ID", "pid")
    monkeypatch.setenv("WHATSAPP_NOTICE_TRANSITION_FACILITATION_2026_NAME", "transition_notice_2026")
    monkeypatch.setenv("WHATSAPP_NOTICE_TRANSITION_FACILITATION_2026_LANG", "en")

    mock_response = type("Resp", (), {
        "status_code": 200,
        "json": lambda self: {"messages": [{"id": "wamid.ABC"}]},
    })()
    with patch("whatsapp_renewal_alerts.requests.post", return_value=mock_response):
        response = client.post("/api/notices/transition_facilitation_2026/send-whatsapp", params={"scheme": "CRS"})
        assert response.status_code == 200
        job_id = response.json()["job_id"]

        import time
        status_response = None
        for _ in range(50):
            status_response = client.get(f"/api/notices/transition_facilitation_2026/send-whatsapp/status/{job_id}")
            if status_response.json()["done"]:
                break
            time.sleep(0.05)

    final = status_response.json()
    assert final["done"] is True
    assert final["sent"] == 1


def test_send_notice_whatsapp_status_returns_404_for_unknown_job():
    response = client.get("/api/notices/transition_facilitation_2026/send-whatsapp/status/does-not-exist")
    assert response.status_code == 404


def test_send_notice_email_starts_job_and_reports_progress(tmp_path, monkeypatch):
    db_path = tmp_path / "clients.db"
    _write_db(db_path, [
        ["CLT001", "Rahul Sharma", "TechCorp", "r@x.com", "919876543210",
         "OSHA", "CRS", "OSHA-1", "01-01-2025", "24-07-2026", "https://x", "ACTIVE"],
    ])
    monkeypatch.setattr(main_module, "DEFAULT_DB_PATH", db_path)
    monkeypatch.setenv("BREVO_API_KEY", "key")
    monkeypatch.setenv("EMAIL_SENDER", "sender@x.com")

    mock_response = type("Resp", (), {
        "status_code": 201,
        "json": lambda self: {"messageId": "brevo-1"},
    })()
    with patch("email_alerts.requests.post", return_value=mock_response):
        response = client.post("/api/notices/transition_facilitation_2026/send-email", params={"scheme": "CRS"})
        assert response.status_code == 200
        job_id = response.json()["job_id"]

        import time
        status_response = None
        for _ in range(50):
            status_response = client.get(f"/api/notices/transition_facilitation_2026/send-email/status/{job_id}")
            if status_response.json()["done"]:
                break
            time.sleep(0.05)

    final = status_response.json()
    assert final["done"] is True
    assert final["sent"] == 1


def test_send_notice_email_status_returns_404_for_unknown_job():
    response = client.get("/api/notices/transition_facilitation_2026/send-email/status/does-not-exist")
    assert response.status_code == 404
