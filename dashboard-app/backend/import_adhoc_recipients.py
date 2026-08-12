"""One-time import: loads a flat phone-number list into the adhoc_recipients
collection for a given ad-hoc notice_id. Not part of any app code path --
run once by hand, per docs/superpowers/specs/2026-08-11-independence-day-whatsapp-broadcast-design.md.

Usage: python import_adhoc_recipients.py "<path to xlsx>" <notice_id>

Expects a sheet named "Numbers" with a "Digits Only" column (matches
Numbers_Only_Deduplicated.xlsx's actual format) and a "Source" column.
Any other sheet (e.g. "Needs Review") is ignored -- this only reads the
"Numbers" sheet."""
import sys

import openpyxl

import db as db_module
from whatsapp_renewal_alerts import normalize_phone


def import_adhoc_recipients(source_path: str, notice_id: str) -> dict:
    wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    ws = wb["Numbers"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    index_map = {str(h).strip().lower(): i for i, h in enumerate(header) if h is not None}

    digits_idx = index_map["digits only"]
    source_idx = index_map.get("source")

    docs = []
    seen = set()
    skipped_blank = 0
    for row in rows:
        raw = row[digits_idx]
        if not raw:
            skipped_blank += 1
            continue
        phone = normalize_phone(raw)
        if not phone or phone in seen:
            continue
        seen.add(phone)
        docs.append({
            "_id": phone,
            "notice_id": notice_id,
            "source": row[source_idx] if source_idx is not None else None,
        })

    wb.close()

    db = db_module.DEFAULT_DB_PATH
    if docs:
        db["adhoc_recipients"].delete_many({"notice_id": notice_id})
        db["adhoc_recipients"].insert_many(docs)

    return {"imported": len(docs), "skipped_blank": skipped_blank}


def demo():
    """Self-check: import logic against an in-memory mongomock database and
    a tiny in-memory workbook, no real file or real database needed."""
    import mongomock
    import db as db_module

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Numbers"
    ws.append(["S.No", "Name", "Phone Number", "Digits Only", "Source"])
    ws.append([1, None, "919876543210", "919876543210", "WhatsApp broadcast list"])
    ws.append([2, None, "0091 98123 45678", "009198123 45678", "WhatsApp broadcast list"])  # exercises the 00-prefix strip
    ws.append([3, None, None, None, "WhatsApp broadcast list"])  # blank row, must be skipped
    wb.save("_demo_adhoc_import.xlsx")

    original_db_path = db_module.DEFAULT_DB_PATH
    db_module.DEFAULT_DB_PATH = mongomock.MongoClient()["demo"]
    try:
        result = import_adhoc_recipients("_demo_adhoc_import.xlsx", "demo_notice")
        assert result == {"imported": 2, "skipped_blank": 1}, result
        stored = list(db_module.DEFAULT_DB_PATH["adhoc_recipients"].find({"notice_id": "demo_notice"}))
        ids = {d["_id"] for d in stored}
        assert ids == {"919876543210", "9198123 45678".replace(" ", "")}, ids
        print("demo() self-check passed:", result)
    finally:
        db_module.DEFAULT_DB_PATH = original_db_path
        import os
        os.remove("_demo_adhoc_import.xlsx")


if __name__ == "__main__":
    if len(sys.argv) == 1:
        demo()
    elif len(sys.argv) == 3:
        source, notice_id = sys.argv[1], sys.argv[2]
        result = import_adhoc_recipients(source, notice_id)
        print(f"Imported {result['imported']} recipients for notice_id={notice_id!r} "
              f"(skipped {result['skipped_blank']} blank rows).")
    else:
        print("Usage: python import_adhoc_recipients.py \"<path to xlsx>\" <notice_id>")
        print("       python import_adhoc_recipients.py   (no args runs the self-check demo)")
        sys.exit(1)
