"""One-off importer: converts a BIS CRS (Compulsory Registration Scheme)
export into clients_certifications.xlsx's single-sheet client roster schema.

Source workbook has one sheet per Indian Standard (e.g. "IS 13252") plus a
"BIS_CRS_Master" sheet that duplicates every per-standard sheet's rows
combined -- the Master sheet is skipped during import to avoid double-
counting every client. There's no expiry/validity date column in this
source format, and no column distinguishing an initial registration from a
renewal either -- Grant Date is all we get. A first-time CRS registration
is valid 2 years; a renewal can be granted for 2-5 years, and this data
can't tell the two apart. ponytail: Expiry Date is computed as Grant Date
+ 5 years everywhere (the maximum real term) rather than the true term per
row, a deliberate choice to avoid the worse failure mode -- understating a
renewed license's real expiry would mean it goes quiet on urgent/critical
alerts while still genuinely valid, whereas overstating it (this choice)
only risks a stale registration slipping past its real 2-year expiry
un-nagged for up to 3 extra years. Upgrade path: if the source ever gains
a term-length or "valid upto" column, read it instead of assuming 5.
License No.
is used for both Client ID and Certification ID, since it's the natural
unique key in this dataset (mirrors import_bis_isi_data.py's use of its own
licence number). A given license number is listed once per Product Category
it covers, so the same license + Indian Standard pair routinely repeats
across several rows -- these collapse to a single roster row (the same
certification, not several); a license number reused under a genuinely
different Indian Standard still gets its own row, suffixed to stay unique.
Status is recomputed from the computed expiry using this project's own
urgency thresholds -- the source "Status" column (Register/Registered/etc.)
means something different and is not used.

Usage: python import_crs.py "<path to source xlsx>"
"""
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

from import_helpers import RowCollector, parse_validity_date, compute_status, header_index_map, get

OUTPUT_HEADERS = [
    "Client ID", "Full Name", "Company", "Email", "Phone (WhatsApp)",
    "Certification Name", "Scheme", "Certification ID", "Issue Date", "Expiry Date",
    "Renewal Link", "Status",
]

MASTER_SHEET_NAME = "bis_crs_master"


def add_years(dt: datetime, years: int) -> datetime:
    """dt with `years` added, falling back to Feb 28 if dt is a Feb 29 whose
    shifted year isn't a leap year."""
    try:
        return dt.replace(year=dt.year + years)
    except ValueError:
        return dt.replace(year=dt.year + years, day=28)


def looks_like_crs_workbook(wb) -> bool:
    """True if any sheet's header row has both a license-no and
    organization-name column. Deliberately different spelling from BIS
    ISI's own detector ("licence no" / "firm name"), so the two formats
    can never be mistaken for each other."""
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            header_row = next(ws.iter_rows(values_only=True))
        except StopIteration:
            continue
        index_map = header_index_map(header_row)
        has_license_no = "license no" in index_map or "license no." in index_map
        if has_license_no and "organization name" in index_map:
            return True
    return False


def import_crs_workbook(wb, out_ws, today=None):
    """Appends converted roster rows from an already-open CRS workbook into
    out_ws. Skips the "BIS_CRS_Master" rollup sheet, which duplicates every
    per-standard sheet's rows combined -- importing it too would count
    every client twice."""
    today = today or datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)

    seen_license_cert_pairs = set()
    seen_license_nos = set()
    stats = {
        "sheets_used": 0, "sheets_empty": 0, "rows_written": 0,
        "rows_skipped_missing_key": 0, "rows_skipped_duplicate_product": 0,
    }

    for sheet_name in wb.sheetnames:
        # Only skip the master rollup sheet when per-standard sheets also
        # exist to import from instead -- if it's the workbook's only sheet,
        # there's nothing to double-count, so it needs to be imported.
        if sheet_name.strip().lower() == MASTER_SHEET_NAME and len(wb.sheetnames) > 1:
            continue

        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            stats["sheets_empty"] += 1
            continue

        index_map = header_index_map(header_row)
        sheet_had_rows = False

        for row in rows_iter:
            if not row or row[0] is None:
                continue
            license_no = get(row, index_map, "license no", "license no.")
            org_name = get(row, index_map, "organization name")
            email = get(row, index_map, "e-mail id", "email id", "e-mail", "email")
            phone = get(row, index_map, "phone no.", "phone no", "phone")
            grant_date_raw = get(row, index_map, "grant date")
            cert_name = get(row, index_map, "indian standard")

            if not license_no or not grant_date_raw:
                stats["rows_skipped_missing_key"] += 1
                continue

            grant_dt = parse_validity_date(grant_date_raw)
            if grant_dt is None:
                stats["rows_skipped_missing_key"] += 1
                continue
            expiry_dt = add_years(grant_dt, 5)

            license_no_clean = str(license_no).strip()
            license_cert_pair = (license_no_clean, cert_name)
            if license_cert_pair in seen_license_cert_pairs:
                # Same license *and* same Indian Standard as a row already
                # written -- this describes the same certification, just
                # listed again for another Product Category (real CRS
                # exports do this routinely). Not a new certification, so
                # not a new row.
                stats["rows_skipped_duplicate_product"] += 1
                continue

            client_id = license_no_clean
            if license_no_clean in seen_license_nos:
                # Same license number but a genuinely different standard --
                # a real second certification sharing an identifier, so it
                # gets its own row.
                client_id = f"{license_no_clean}-{cert_name}"
            seen_license_cert_pairs.add(license_cert_pair)
            seen_license_nos.add(license_no_clean)

            out_ws.append([
                client_id,
                org_name,
                org_name,
                email,
                phone,
                cert_name,
                "CRS",
                license_no_clean,
                grant_dt.strftime("%d-%m-%Y"),
                expiry_dt.strftime("%d-%m-%Y"),
                None,
                compute_status(expiry_dt, today),
            ])
            stats["rows_written"] += 1
            sheet_had_rows = True

        if sheet_had_rows:
            stats["sheets_used"] += 1
        else:
            stats["sheets_empty"] += 1

    return stats


def import_crs(source_path: str, output_path: str):
    wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.append(OUTPUT_HEADERS)

    stats = import_crs_workbook(wb, out_ws)

    wb.close()
    out_wb.save(output_path)
    return stats


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python import_crs.py \"<path to source xlsx>\"")
        sys.exit(1)

    source = sys.argv[1]
    output = str(Path(__file__).parent.parent.parent / "data" / "clients_certifications.xlsx")
    result = import_crs(source, output)
    print("Sheets with data used:", result["sheets_used"])
    print("Sheets empty/skipped:", result["sheets_empty"])
    print("Rows written:", result["rows_written"])
    print("Rows skipped (missing license no / grant date):", result["rows_skipped_missing_key"])
    print("Rows skipped (duplicate license + standard, different product category):",
          result["rows_skipped_duplicate_product"])
    print("Saved to:", output)
