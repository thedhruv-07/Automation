"""FastAPI backend for the Absolute Veritas React dashboard."""
from pathlib import Path

BACKEND_DIR = Path(__file__).parent
REPO_ROOT = BACKEND_DIR.parent.parent

import base64
import io
import os
import sqlite3
import threading
import uuid

import openpyxl
from dotenv import load_dotenv
from fastapi import FastAPI, Form, HTTPException, File, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse

from datetime import datetime

from db import (  # noqa: E402
    DEFAULT_DB_PATH, get_clients_page, get_stats, export_clients_rows,
    upsert_clients, find_client_by_id, load_sent_log, save_sent_log,
    is_already_sent, load_email_sent_log, save_email_sent_log, is_email_already_sent,
    get_eligible_count, get_notice_eligible_count, get_broadcast_clients_page,
)
from whatsapp_renewal_alerts import (  # noqa: E402
    ALERT_STATUSES, dedup_key, filter_alertable, normalize_phone,
    send_one_alert, run,
)
from email_alerts import (  # noqa: E402
    send_email_via_brevo, send_one_email_alert, run_email_alerts,
)
from email_template import build_email_html  # noqa: E402
from import_helpers import RowCollector, REQUIRED_HEADERS  # noqa: E402
from import_formats import IMPORT_FORMATS, FORMAT_DISPLAY_NAMES  # noqa: E402
from scheme_templates import get_email_content  # noqa: E402
from notices import list_notices, get_notice_module  # noqa: E402
from notice_sender import send_notice_whatsapp, send_notice_email  # noqa: E402

load_dotenv(REPO_ROOT / ".env")

EMAIL_DATE_FORMATS = ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y")

LOGO_PATH = REPO_ROOT / "dashboard-app" / "frontend" / "public" / "company-logo.png"


def _logo_data_uri() -> str:
    if not LOGO_PATH.exists():
        return ""
    return "data:image/png;base64," + base64.b64encode(LOGO_PATH.read_bytes()).decode("ascii")


def _today_str() -> str:
    return datetime.now().strftime("%Y-%m-%d")


def _parse_expiry(value) -> datetime:
    if isinstance(value, datetime):
        return value
    for fmt in EMAIL_DATE_FORMATS:
        try:
            return datetime.strptime(str(value).strip(), fmt)
        except ValueError:
            continue
    raise ValueError(f"Unrecognized date format: {value!r}")


app = FastAPI(title="Absolute Veritas Renewal Dashboard API")

_send_lock = threading.Lock()
_pending_sends: set[str] = set()

_bulk_in_progress = False

_email_send_lock = threading.Lock()
_pending_email_sends: set[str] = set()

_email_bulk_in_progress = False

# Mirrors the hardcoded day cutoffs in cert_automation.py (CRITICAL <= 7 days,
# URGENT <= 30 days). That script is intentionally untouched by this project,
# so these are documented here for display only, not read live from it.
CERT_STATUS_THRESHOLDS = {
    "critical_days": 7,
    "urgent_days": 30,
}


def _resolve_allowed_origins() -> list[str]:
    origins = ["http://localhost:5173", "http://127.0.0.1:5173"]
    extra = os.environ.get("DASHBOARD_ALLOWED_ORIGIN")
    if extra:
        origins.append(extra)
    return origins


app.add_middleware(
    CORSMiddleware,
    allow_origins=_resolve_allowed_origins(),
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/api/health")
def health():
    return {"status": "ok"}


@app.get("/api/clients")
def get_clients(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=500),
    status: str = "ALL", cert_type: str = "ALL", scheme: str = "ALL",
    expiry_before: str = "", search: str = "", sort_key: str = "", sort_dir: str = "asc",
):
    today = _today_str()
    rows, total = get_clients_page(
        DEFAULT_DB_PATH, page=page, page_size=page_size,
        status=status, cert_type=cert_type, scheme=scheme, expiry_before=expiry_before or None,
        search=search or None, sort_key=sort_key or None, sort_dir=sort_dir.lower(),
    )
    result = []
    for rec in rows:
        if rec["status"] in ALERT_STATUSES:
            alert_sent_today = is_already_sent(DEFAULT_DB_PATH, rec["client_id"], rec["status"], today)
        else:
            alert_sent_today = None
        result.append({**rec, "alert_sent_today": alert_sent_today})
    return {"rows": result, "total": total, "page": page, "page_size": page_size}


@app.get("/api/stats")
def stats():
    return get_stats(DEFAULT_DB_PATH, _today_str())


@app.get("/api/eligible-count")
def eligible_count(
    status: str = "", cert_type: str = "", expiry_before: str = "", search: str = "", scheme: str = "",
):
    today = _today_str()
    return {
        "whatsapp": get_eligible_count(
            DEFAULT_DB_PATH, today, "whatsapp",
            status=status or None, cert_type=cert_type or None, expiry_before=expiry_before or None,
            search=search or None, scheme=scheme or None,
        ),
        "email": get_eligible_count(
            DEFAULT_DB_PATH, today, "email",
            status=status or None, cert_type=cert_type or None, expiry_before=expiry_before or None,
            search=search or None, scheme=scheme or None,
        ),
    }


def _csv_escape(value) -> str:
    text = str(value if value is not None else "")
    if text.startswith(("=", "+", "-", "@")):
        text = "'" + text
    if any(c in text for c in ('"', ",", "\n")):
        return '"' + text.replace('"', '""') + '"'
    return text


@app.get("/api/clients/export")
def export_clients(
    status: str = "ALL", cert_type: str = "ALL", expiry_before: str = "", search: str = "", scheme: str = "ALL",
):
    def generate():
        yield ",".join(_csv_escape(h) for h in REQUIRED_HEADERS) + "\n"
        for rec in export_clients_rows(
            DEFAULT_DB_PATH, status=status, cert_type=cert_type, scheme=scheme,
            expiry_before=expiry_before or None, search=search or None,
        ):
            values = [
                rec["client_id"], rec["name"], rec["company"], rec["email"], rec["phone"],
                rec["cert_name"], rec["scheme"], rec["cert_id"], rec["issue_date"], rec["expiry_date"],
                rec["renewal_link"], rec["status"],
            ]
            yield ",".join(_csv_escape(v) for v in values) + "\n"

    return StreamingResponse(
        generate(), media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=clients_export.csv"},
    )


@app.get("/api/email-preview/{client_id}")
def email_preview(client_id: str):
    record = find_client_by_id(DEFAULT_DB_PATH, client_id)
    if record is None:
        raise HTTPException(status_code=404, detail=f"Unknown client_id: {client_id}")

    expiry_dt = _parse_expiry(record["expiry_date"])
    days_left = (expiry_dt - datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)).days
    rec = {
        **record,
        "days_left": days_left,
        "expiry_formatted": expiry_dt.strftime("%d %B %Y"),
    }
    subject_template, intro_text = get_email_content(record["scheme"])
    html = build_email_html(
        rec,
        org_name="Absolute Veritas",
        org_website="",
        org_contact="",
        org_email="cs@absoluteveritas.com",
        logo_src=_logo_data_uri(),
        intro_text=intro_text,
    )
    subject = subject_template.format(cert_name=record["cert_name"], company=record["company"])
    return {"subject": subject, "html": html}


@app.get("/api/settings-info")
def settings_info():
    return {
        "template_name": os.environ.get("WHATSAPP_TEMPLATE_NAME", "cert_renewal_alert"),
        "template_lang": os.environ.get("WHATSAPP_TEMPLATE_LANG", "en"),
        "phone_number_id": os.environ.get("PHONE_NUMBER_ID", ""),
        "scheduled_run_time": "09:30 (Asia/Kolkata)",
        "critical_days": CERT_STATUS_THRESHOLDS["critical_days"],
        "urgent_days": CERT_STATUS_THRESHOLDS["urgent_days"],
    }


@app.get("/api/message-log")
def message_log():
    sent_log = load_sent_log(DEFAULT_DB_PATH)
    entries = []
    for key, info in sent_log.items():
        client_id, status_tier, _date = key.split("|", 2)
        found = find_client_by_id(DEFAULT_DB_PATH, client_id) or {}
        entries.append({
            "client_id": client_id,
            "name": found.get("name", "Unknown"),
            "company": found.get("company", ""),
            "cert_name": found.get("cert_name", ""),
            "status_tier": status_tier,
            "phone": info.get("phone"),
            "message_id": info.get("message_id"),
            "sent_at": info.get("sent_at"),
        })
    entries.sort(key=lambda e: e["sent_at"] or "", reverse=True)
    return entries


@app.post("/api/send/{client_id}")
def send_alert(client_id: str):
    today = _today_str()
    record = find_client_by_id(DEFAULT_DB_PATH, client_id)
    if record is None:
        raise HTTPException(status_code=404, detail=f"Unknown client_id: {client_id}")
    if record["status"] not in ALERT_STATUSES:
        raise HTTPException(
            status_code=400,
            detail=f"Status {record['status']} is not alert-eligible",
        )

    sent_log = load_sent_log(DEFAULT_DB_PATH)
    key = dedup_key(record["client_id"], record["status"], today)
    if key in sent_log:
        raise HTTPException(
            status_code=409,
            detail="Alert already sent today for this client/status",
        )

    with _send_lock:
        if client_id in _pending_sends:
            raise HTTPException(
                status_code=409,
                detail="A send for this client is already in progress",
            )
        if _bulk_in_progress:
            raise HTTPException(
                status_code=409,
                detail="A bulk send is in progress; try again after it completes",
            )
        _pending_sends.add(client_id)

    try:
        token = os.environ["WHATSAPP_TOKEN"]
        phone_number_id = os.environ["PHONE_NUMBER_ID"]
        test_number = os.environ.get("DASHBOARD_TEST_NUMBER") or None

        result = send_one_alert(
            record, sent_log, today, token, phone_number_id, to_phone_override=test_number,
        )

        if result["action"] == "sent":
            if not test_number:
                save_sent_log(DEFAULT_DB_PATH, sent_log)
            return {"status": "sent", "message_id": result["message_id"]}
        if result["action"] == "skipped_duplicate":
            raise HTTPException(
                status_code=409,
                detail="Alert already sent today for this client/status",
            )
        if result["action"] == "skipped_no_template":
            raise HTTPException(
                status_code=400,
                detail=f"No WhatsApp template configured for scheme {record['scheme']!r} yet.",
            )
        raise HTTPException(status_code=502, detail=result.get("error", "Unknown error"))
    finally:
        with _send_lock:
            _pending_sends.discard(client_id)


_send_all_jobs: dict[str, dict] = {}


def _run_send_all_job(
    job_id, token, phone_number_id, test_number,
    status=None, cert_type=None, expiry_before=None, search=None, scheme=None,
):
    def progress(result, total):
        job = _send_all_jobs[job_id]
        job["total"] = total
        if result["action"] == "sent":
            job["sent"] += 1
        elif result["action"] == "skipped_duplicate":
            job["skipped"] += 1
        elif result["action"] == "skipped_no_template":
            job["skipped_no_template"] += 1
        elif result["action"] == "failed":
            job["failed"] += 1

    try:
        run(
            DEFAULT_DB_PATH, token, phone_number_id,
            dry_run=False, test_number=test_number, on_progress=progress,
            status=status, cert_type=cert_type, expiry_before=expiry_before, search=search, scheme=scheme,
        )
    except Exception as exc:
        # Without this, an exception here (locked DB, unexpected error mid-
        # send-loop) only propagates into this daemon thread -- printed to
        # stderr, invisible to the running app -- while the job dict is left
        # at done=True with whatever counts happened to accumulate before
        # the crash (often 0/0/0), indistinguishable from a real "nothing to
        # do" success. Recording it here lets polling clients tell the two
        # apart.
        _send_all_jobs[job_id]["error"] = str(exc)
    finally:
        _send_all_jobs[job_id]["done"] = True
        global _bulk_in_progress
        with _send_lock:
            _bulk_in_progress = False


@app.post("/api/send-all")
def send_all_alerts(
    status: str = "", cert_type: str = "", expiry_before: str = "", search: str = "", scheme: str = "",
):
    global _bulk_in_progress
    with _send_lock:
        if _bulk_in_progress:
            raise HTTPException(status_code=409, detail="A bulk send is already in progress")
        if _pending_sends:
            raise HTTPException(
                status_code=409,
                detail="One or more per-client sends are in progress; try again shortly",
            )
        _bulk_in_progress = True

    try:
        token = os.environ["WHATSAPP_TOKEN"]
        phone_number_id = os.environ["PHONE_NUMBER_ID"]
        test_number = os.environ.get("DASHBOARD_TEST_NUMBER") or None

        job_id = str(uuid.uuid4())
        _send_all_jobs[job_id] = {
            "total": 0, "sent": 0, "skipped": 0, "skipped_no_template": 0, "failed": 0,
            "done": False, "error": None,
        }
        thread = threading.Thread(
            target=_run_send_all_job,
            args=(
                job_id, token, phone_number_id, test_number,
                status or None, cert_type or None, expiry_before or None, search or None, scheme or None,
            ),
            daemon=True,
        )
        thread.start()
        return {"job_id": job_id}
    except Exception:
        with _send_lock:
            _bulk_in_progress = False
        raise HTTPException(status_code=500, detail="Server is not configured to send WhatsApp messages")


@app.get("/api/send-all/status/{job_id}")
def send_all_status(job_id: str):
    job = _send_all_jobs.get(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="Unknown job_id")
    return job


@app.post("/api/send-email/{client_id}")
def send_email(client_id: str):
    today = _today_str()
    record = find_client_by_id(DEFAULT_DB_PATH, client_id)
    if record is None:
        raise HTTPException(status_code=404, detail=f"Unknown client_id: {client_id}")
    if record["status"] not in ALERT_STATUSES:
        raise HTTPException(
            status_code=400,
            detail=f"Status {record['status']} is not alert-eligible",
        )
    if not record.get("email") or "@" not in record["email"]:
        raise HTTPException(
            status_code=400,
            detail="This client has no valid email on file",
        )

    sent_log = load_email_sent_log(DEFAULT_DB_PATH)
    key = dedup_key(record["client_id"], record["status"], today)
    if key in sent_log:
        raise HTTPException(
            status_code=409,
            detail="Email already sent today for this client/status",
        )

    with _email_send_lock:
        if client_id in _pending_email_sends:
            raise HTTPException(
                status_code=409,
                detail="An email send for this client is already in progress",
            )
        if _email_bulk_in_progress:
            raise HTTPException(
                status_code=409,
                detail="A bulk email send is in progress; try again after it completes",
            )
        _pending_email_sends.add(client_id)

    try:
        brevo_api_key = os.environ["BREVO_API_KEY"]
        email_sender = os.environ["EMAIL_SENDER"]
        test_email = os.environ.get("DASHBOARD_TEST_EMAIL") or None

        result = send_one_email_alert(
            record, sent_log, today, brevo_api_key, email_sender, "Absolute Veritas",
            to_email_override=test_email,
        )

        if result["action"] == "sent":
            if not test_email:
                save_email_sent_log(DEFAULT_DB_PATH, sent_log)
            return {"status": "sent", "message_id": result["message_id"]}
        if result["action"] == "skipped_duplicate":
            raise HTTPException(
                status_code=409,
                detail="Email already sent today for this client/status",
            )
        raise HTTPException(status_code=502, detail=result.get("error", "Unknown error"))
    finally:
        with _email_send_lock:
            _pending_email_sends.discard(client_id)


_send_all_email_jobs: dict[str, dict] = {}


def _run_send_all_email_job(
    job_id, brevo_api_key, email_sender, test_email,
    status=None, cert_type=None, expiry_before=None, search=None, scheme=None,
):
    def progress(result, total):
        job = _send_all_email_jobs[job_id]
        job["total"] = total
        if result["action"] == "sent":
            job["sent"] += 1
        elif result["action"] == "skipped_duplicate":
            job["skipped"] += 1
        elif result["action"] == "skipped_no_email":
            job["skipped_no_email"] += 1
        elif result["action"] == "failed":
            job["failed"] += 1

    try:
        run_email_alerts(
            DEFAULT_DB_PATH, brevo_api_key, email_sender, "Absolute Veritas",
            dry_run=False, test_email=test_email, on_progress=progress,
            status=status, cert_type=cert_type, expiry_before=expiry_before,
            search=search, scheme=scheme,
        )
    except Exception as exc:
        _send_all_email_jobs[job_id]["error"] = str(exc)
    finally:
        _send_all_email_jobs[job_id]["done"] = True
        global _email_bulk_in_progress
        with _email_send_lock:
            _email_bulk_in_progress = False


@app.post("/api/send-all-emails")
def send_all_emails(
    status: str = "", cert_type: str = "", expiry_before: str = "", search: str = "", scheme: str = "",
):
    global _email_bulk_in_progress
    with _email_send_lock:
        if _email_bulk_in_progress:
            raise HTTPException(status_code=409, detail="A bulk email send is already in progress")
        if _pending_email_sends:
            raise HTTPException(
                status_code=409,
                detail="One or more per-client email sends are in progress; try again shortly",
            )
        _email_bulk_in_progress = True

    try:
        brevo_api_key = os.environ["BREVO_API_KEY"]
        email_sender = os.environ["EMAIL_SENDER"]
        test_email = os.environ.get("DASHBOARD_TEST_EMAIL") or None

        job_id = str(uuid.uuid4())
        _send_all_email_jobs[job_id] = {
            "total": 0, "sent": 0, "skipped": 0, "skipped_no_email": 0, "failed": 0,
            "done": False, "error": None,
        }
        thread = threading.Thread(
            target=_run_send_all_email_job,
            args=(
                job_id, brevo_api_key, email_sender, test_email,
                status or None, cert_type or None, expiry_before or None, search or None, scheme or None,
            ),
            daemon=True,
        )
        thread.start()
        return {"job_id": job_id}
    except Exception:
        with _email_send_lock:
            _email_bulk_in_progress = False
        raise HTTPException(status_code=500, detail="Server is not configured to send emails")


@app.get("/api/send-all-emails/status/{job_id}")
def send_all_emails_status(job_id: str):
    job = _send_all_email_jobs.get(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="Unknown job_id")
    return job


@app.get("/api/notices")
def notices_list():
    return list_notices()


@app.get("/api/notices/{notice_id}/preview")
def notice_preview(notice_id: str):
    """Renders the notice's email using placeholder client data -- unlike
    /api/email-preview/{client_id} (a specific client's renewal email), a
    notice isn't about any individual client's own record, so there's
    nothing client-specific to preview against."""
    module = get_notice_module(notice_id)
    if module is None:
        raise HTTPException(status_code=404, detail=f"Unknown notice_id: {notice_id}")
    placeholder = {"client_id": "SAMPLE", "name": "Sample Client", "company": "Sample Company"}
    html = module.build_email_html(placeholder, "Absolute Veritas")
    return {"subject": module.EMAIL_SUBJECT, "html": html}


@app.get("/api/notices/{notice_id}/eligible-count")
def notice_eligible_count(
    notice_id: str, status: str = "", cert_type: str = "", expiry_before: str = "",
    search: str = "", scheme: str = "",
):
    if get_notice_module(notice_id) is None:
        raise HTTPException(status_code=404, detail=f"Unknown notice_id: {notice_id}")
    return {
        "whatsapp": get_notice_eligible_count(
            DEFAULT_DB_PATH, notice_id, "whatsapp",
            status=status or None, cert_type=cert_type or None, expiry_before=expiry_before or None,
            search=search or None, scheme=scheme or None,
        ),
        "email": get_notice_eligible_count(
            DEFAULT_DB_PATH, notice_id, "email",
            status=status or None, cert_type=cert_type or None, expiry_before=expiry_before or None,
            search=search or None, scheme=scheme or None,
        ),
    }


@app.get("/api/notices/{notice_id}/clients")
def notice_clients(
    notice_id: str,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=8, ge=1, le=500),
    status: str = "", cert_type: str = "", expiry_before: str = "", search: str = "", scheme: str = "",
):
    if get_notice_module(notice_id) is None:
        raise HTTPException(status_code=404, detail=f"Unknown notice_id: {notice_id}")
    rows, total = get_broadcast_clients_page(
        DEFAULT_DB_PATH, notice_id, page=page, page_size=page_size,
        status=status or None, cert_type=cert_type or None, expiry_before=expiry_before or None,
        search=search or None, scheme=scheme or None,
    )
    return {"rows": rows, "total": total, "page": page, "page_size": page_size}


_send_notice_jobs: dict[str, dict] = {}


def _run_send_notice_whatsapp_job(
    job_id, notice_id, token, phone_number_id, test_number,
    status=None, cert_type=None, expiry_before=None, search=None, scheme=None,
):
    def progress(result, total):
        job = _send_notice_jobs[job_id]
        job["total"] = total
        if result["action"] == "sent":
            job["sent"] += 1
        elif result["action"] == "skipped_duplicate":
            job["skipped"] += 1
        elif result["action"] == "skipped_no_template":
            job["skipped_no_template"] += 1
        elif result["action"] == "failed":
            job["failed"] += 1

    try:
        send_notice_whatsapp(
            DEFAULT_DB_PATH, notice_id, token, phone_number_id,
            dry_run=False, test_number=test_number, on_progress=progress,
            status=status, cert_type=cert_type, expiry_before=expiry_before, search=search, scheme=scheme,
        )
    except Exception as exc:
        _send_notice_jobs[job_id]["error"] = str(exc)
    finally:
        _send_notice_jobs[job_id]["done"] = True


@app.post("/api/notices/{notice_id}/send-whatsapp")
def send_notice_whatsapp_endpoint(
    notice_id: str, status: str = "", cert_type: str = "", expiry_before: str = "",
    search: str = "", scheme: str = "",
):
    if get_notice_module(notice_id) is None:
        raise HTTPException(status_code=404, detail=f"Unknown notice_id: {notice_id}")
    try:
        token = os.environ["WHATSAPP_TOKEN"]
        phone_number_id = os.environ["PHONE_NUMBER_ID"]
        test_number = os.environ.get("DASHBOARD_TEST_NUMBER") or None

        job_id = str(uuid.uuid4())
        _send_notice_jobs[job_id] = {
            "total": 0, "sent": 0, "skipped": 0, "skipped_no_template": 0, "failed": 0,
            "done": False, "error": None,
        }
        thread = threading.Thread(
            target=_run_send_notice_whatsapp_job,
            args=(
                job_id, notice_id, token, phone_number_id, test_number,
                status or None, cert_type or None, expiry_before or None, search or None, scheme or None,
            ),
            daemon=True,
        )
        thread.start()
        return {"job_id": job_id}
    except Exception:
        raise HTTPException(status_code=500, detail="Server is not configured to send WhatsApp messages")


@app.get("/api/notices/{notice_id}/send-whatsapp/status/{job_id}")
def send_notice_whatsapp_status(notice_id: str, job_id: str):
    job = _send_notice_jobs.get(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="Unknown job_id")
    return job


_send_notice_email_jobs: dict[str, dict] = {}


def _run_send_notice_email_job(
    job_id, notice_id, brevo_api_key, email_sender, test_email,
    status=None, cert_type=None, expiry_before=None, search=None, scheme=None,
):
    def progress(result, total):
        job = _send_notice_email_jobs[job_id]
        job["total"] = total
        if result["action"] == "sent":
            job["sent"] += 1
        elif result["action"] == "skipped_duplicate":
            job["skipped"] += 1
        elif result["action"] == "skipped_no_email":
            job["skipped_no_email"] += 1
        elif result["action"] == "failed":
            job["failed"] += 1

    try:
        send_notice_email(
            DEFAULT_DB_PATH, notice_id, brevo_api_key, email_sender, "Absolute Veritas",
            dry_run=False, test_email=test_email, on_progress=progress,
            status=status, cert_type=cert_type, expiry_before=expiry_before, search=search, scheme=scheme,
        )
    except Exception as exc:
        _send_notice_email_jobs[job_id]["error"] = str(exc)
    finally:
        _send_notice_email_jobs[job_id]["done"] = True


@app.post("/api/notices/{notice_id}/send-email")
def send_notice_email_endpoint(
    notice_id: str, status: str = "", cert_type: str = "", expiry_before: str = "",
    search: str = "", scheme: str = "",
):
    if get_notice_module(notice_id) is None:
        raise HTTPException(status_code=404, detail=f"Unknown notice_id: {notice_id}")
    try:
        brevo_api_key = os.environ["BREVO_API_KEY"]
        email_sender = os.environ["EMAIL_SENDER"]
        test_email = os.environ.get("DASHBOARD_TEST_EMAIL") or None

        job_id = str(uuid.uuid4())
        _send_notice_email_jobs[job_id] = {
            "total": 0, "sent": 0, "skipped": 0, "skipped_no_email": 0, "failed": 0,
            "done": False, "error": None,
        }
        thread = threading.Thread(
            target=_run_send_notice_email_job,
            args=(
                job_id, notice_id, brevo_api_key, email_sender, test_email,
                status or None, cert_type or None, expiry_before or None, search or None, scheme or None,
            ),
            daemon=True,
        )
        thread.start()
        return {"job_id": job_id}
    except Exception:
        raise HTTPException(status_code=500, detail="Server is not configured to send emails")


@app.get("/api/notices/{notice_id}/send-email/status/{job_id}")
def send_notice_email_status(notice_id: str, job_id: str):
    job = _send_notice_email_jobs.get(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="Unknown job_id")
    return job


@app.get("/api/client-template")
def client_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(REQUIRED_HEADERS)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=clients_certifications_template.xlsx"},
    )


def _upsert_clients_or_400(rows, mode):
    """Wraps upsert_clients() so a dirty row (e.g. a blank required field
    like name, which violates the `name TEXT NOT NULL` constraint) surfaces
    to the admin as a clear 400 instead of an opaque 500. db.py's
    upsert_clients() rolls back the whole batch on any such failure, so by
    the time this exception reaches here nothing from the batch was
    written -- the 400 accurately reflects "nothing was saved"."""
    try:
        return upsert_clients(DEFAULT_DB_PATH, rows, mode=mode)
    except sqlite3.IntegrityError as exc:
        raise HTTPException(
            status_code=400,
            detail=(
                "Upload contains invalid data (e.g. a blank required field like "
                f"name) and was not saved: {exc}"
            ),
        )


@app.post("/api/upload-clients")
async def upload_clients(file: UploadFile = File(...), import_format: str = Form(...)):
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="File must be an .xlsx spreadsheet")

    format_entry = next((f for f in IMPORT_FORMATS if f[0] == import_format), None)
    if format_entry is None:
        raise HTTPException(status_code=400, detail=f"Unknown format {import_format!r}.")
    _, detector, importer, expected_columns = format_entry

    contents = await file.read()
    tmp_path = DEFAULT_DB_PATH.parent / "_upload_tmp.xlsx"
    tmp_path.write_bytes(contents)

    try:
        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
        active_title = wb.active.title
        try:
            header_row = next(wb.active.iter_rows(values_only=True))
        except StopIteration:
            header_row = None
    except Exception:
        tmp_path.unlink(missing_ok=True)
        raise HTTPException(status_code=400, detail="Could not read the uploaded file as a valid .xlsx spreadsheet")

    if header_row is None:
        wb.close()
        tmp_path.unlink(missing_ok=True)
        raise HTTPException(
            status_code=400,
            detail=(
                f"The active sheet ('{active_title}') has no rows. If this file has "
                "multiple sheets, make sure the one with your client data is the sheet "
                "selected/visible when the file was last saved."
            ),
        )

    if not detector(wb):
        wb.close()
        tmp_path.unlink(missing_ok=True)
        display_name = FORMAT_DISPLAY_NAMES.get(import_format, import_format)
        raise HTTPException(
            status_code=400,
            detail=f"This doesn't look like a valid {display_name} export — expected {expected_columns}.",
        )

    collector = RowCollector()
    format_stats = importer(wb, collector)
    wb.close()
    tmp_path.unlink(missing_ok=True)

    if format_stats["rows_written"] == 0:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Recognized this as a {import_format} file, but no rows had both a "
                "required identifier and a validity date to import."
            ),
        )

    stats = _upsert_clients_or_400(collector.rows, mode="replace")
    return {"status": "ok", "row_count": stats["row_count"], "format": import_format, "stats": format_stats}


@app.post("/api/merge-clients")
async def merge_clients(file: UploadFile = File(...), import_format: str = Form(...)):
    """Adds rows from the uploaded spreadsheet to the existing roster instead
    of replacing it. Client IDs already present in the roster are left
    untouched — only genuinely new Client IDs get appended."""
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="File must be an .xlsx spreadsheet")

    format_entry = next((f for f in IMPORT_FORMATS if f[0] == import_format), None)
    if format_entry is None:
        raise HTTPException(status_code=400, detail=f"Unknown format {import_format!r}.")
    _, detector, importer, expected_columns = format_entry

    contents = await file.read()
    tmp_path = DEFAULT_DB_PATH.parent / "_merge_upload_tmp.xlsx"
    tmp_path.write_bytes(contents)

    try:
        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
        active_title = wb.active.title
        try:
            header_row = next(wb.active.iter_rows(values_only=True))
        except StopIteration:
            header_row = None
    except Exception:
        tmp_path.unlink(missing_ok=True)
        raise HTTPException(status_code=400, detail="Could not read the uploaded file as a valid .xlsx spreadsheet")

    if header_row is None:
        wb.close()
        tmp_path.unlink(missing_ok=True)
        raise HTTPException(
            status_code=400,
            detail=(
                f"The active sheet ('{active_title}') has no rows. If this file has "
                "multiple sheets, make sure the one with your client data is the sheet "
                "selected/visible when the file was last saved."
            ),
        )

    if not detector(wb):
        wb.close()
        tmp_path.unlink(missing_ok=True)
        display_name = FORMAT_DISPLAY_NAMES.get(import_format, import_format)
        raise HTTPException(
            status_code=400,
            detail=f"This doesn't look like a valid {display_name} export — expected {expected_columns}.",
        )

    collector = RowCollector()
    format_stats = importer(wb, collector)
    wb.close()
    tmp_path.unlink(missing_ok=True)

    if format_stats["rows_written"] == 0:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Recognized this as a {import_format} file, but no rows had both a "
                "required identifier and a validity date to import."
            ),
        )

    stats = _upsert_clients_or_400(collector.rows, mode="merge")
    return {
        "status": "ok", "row_count": stats["row_count"], "added": stats["added"],
        "skipped_duplicates": stats["skipped_duplicates"], "format": import_format, "stats": format_stats,
    }


from fastapi.staticfiles import StaticFiles  # noqa: E402

FRONTEND_DIST = REPO_ROOT / "dashboard-app" / "frontend" / "dist"
if FRONTEND_DIST.exists():
    app.mount("/", StaticFiles(directory=str(FRONTEND_DIST), html=True), name="frontend")
else:
    print(f"Frontend not built — run 'npm run build' in {FRONTEND_DIST.parent} before starting the server.")
